from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
from database import init_db, get_db_connection, DATABASE, update_database_schema
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
from auth import User, has_permission
from utils import normalize_number, normalize_date, format_excel_number
from excel_import import ExcelImporter
import shutil
import zipfile
import hashlib
import json
import io
import re as _re
from flask import send_from_directory
from flask import jsonify, request
from flask_login import login_required, current_user



app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'  # Change this!

# Upload configuration
UPLOAD_FOLDER = 'data/uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.get(user_id)

# Initialize folders and database on first run
os.makedirs('data', exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs('templates/pages', exist_ok=True)  # Create pages folder

if not os.path.exists(DATABASE):
    init_db()
else:
    update_database_schema()

# Ensure currency column exists on order_lines (order_lines is source of truth for currency)
def _ensure_currency_column():
    try:
        conn = get_db_connection()
        # Add to order_lines (source of truth per user requirement)
        ol_cols = [r[1] for r in conn.execute("PRAGMA table_info(order_lines)").fetchall()]
        if 'currency' not in ol_cols:
            conn.execute("ALTER TABLE order_lines ADD COLUMN currency TEXT DEFAULT 'EUR'")
            conn.commit()
            print("[schema] Added order_lines.currency column")
        # Also add to orders for header-level fast retrieval
        o_cols = [r[1] for r in conn.execute("PRAGMA table_info(orders)").fetchall()]
        if 'currency' not in o_cols:
            conn.execute("ALTER TABLE orders ADD COLUMN currency TEXT DEFAULT 'EUR'")
            conn.commit()
            print("[schema] Added orders.currency column")
        conn.close()
    except Exception as e:
        print(f"[schema] currency column check: {e}")
_ensure_currency_column()

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ============== AUTHENTICATION ROUTES ==============

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        language = request.form.get('language', 'en')  # Get language from hidden field
        
        user = User.authenticate(username, password)
        if user:
            # Save language to session
            session['language'] = language
            
            # Update user's language preference in database
            user.update_language(language)
            
            # Log the user in
            login_user(user)
            
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
            return render_template('login.html', error='Invalid username or password')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/favicon.ico')
def favicon():
    """Serve favicon from static/images/logo.ico"""
    return send_from_directory(
        os.path.join(app.root_path, 'static', 'images'),
        'logo.ico',
        mimetype='image/vnd.microsoft.icon'
    )    

# ============== PAGE ROUTES (Dynamic Loading) ==============

@app.route('/page/<page_name>')
@login_required
def load_page(page_name):
    """Load individual page HTML"""
    try:
        # Security: whitelist allowed pages
        allowed_pages = [
            'backup', 'restore', 'mission-details', 'projects', 'end-users', 'third-parties',
            'user-management',
            'order-generation', 'orders-followup', 'order-details', 'back-orders',
            'cargo-reception', 'cargo-followup', 'upload-cargo',
            'dispatch-parcels', 'dispatch-item', 'receive-parcel', 'receive-item', 'parcel-followup',
            'stock-availability', 'stock-card', 'parcel-tracing', 'change-location',
            'donations', 'losses', 'sleeping-stock', 'expiry-report',
            # Movements & Inventory
            'movements-in', 'movements-out', 'reports', 'inventory',
            'reception-report',
        ]
        
        if page_name not in allowed_pages:
            return "<div class='page-content active'><div class='coming-soon'>Page not found</div></div>", 404
        
        # Convert page-name to page_name.html
        template_name = f"pages/{page_name.replace('-', '_')}.html"
        
        return render_template(template_name)
    
    except Exception as e:
        # Return coming soon page if template doesn't exist
        return f"""
        <div class='page-content active'>
            <div class='page-header'>
                <h2>{page_name.replace('-', ' ').title()}</h2>
            </div>
            <div class='coming-soon'>This page is under development. Coming soon!</div>
        </div>
        """, 200


@app.route('/page/end-users')
@login_required
def page_end_users():
    """End users management page"""
    return render_template('pages/end_users.html')        

# ============== API ROUTES ==============

@app.route('/api/user-language', methods=['GET'])
@login_required
def get_user_language():
    """Get current user's language preference"""
    language = session.get('language', current_user.language if hasattr(current_user, 'language') else 'en')
    return jsonify({'language': language})

@app.route('/api/set-language', methods=['POST'])
@login_required
def set_language():
    """Set user's language preference"""
    data = request.json
    language = data.get('language', 'en')
    
    # Save to session
    session['language'] = language
    
    # Update user in database
    current_user.update_language(language)
    
    return jsonify({'success': True, 'language': language})

@app.route('/api/user-role', methods=['GET'])
@login_required
def get_user_role():
    """Get current user's role"""
    return jsonify({'role': current_user.role})

@app.route('/api/items', methods=['GET', 'POST'])
@login_required
def manage_items():
    """Get all items or add new item"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        if not has_permission(current_user, 'manage_items'):
            return jsonify({'success': False, 'message': 'Permission denied'}), 403
        
        data = request.json
        barcode = data.get('barcode')
        name = data.get('name')
        quantity = normalize_number(data.get('quantity', 0))
        location = data.get('location', '')
        
        try:
            conn.execute(
                'INSERT INTO items (barcode, name, quantity, location) VALUES (?, ?, ?, ?)',
                (barcode, name, quantity, location)
            )
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Item added successfully'})
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'success': False, 'message': 'Barcode already exists'}), 400
    
    else:  # GET
        items = conn.execute('SELECT * FROM items ORDER BY created_at DESC').fetchall()
        conn.close()
        return jsonify([dict(item) for item in items])

@app.route('/api/items/<barcode>', methods=['GET'])
@login_required
def get_item_by_barcode(barcode):
    """Get item by barcode for validation"""
    conn = get_db_connection()
    item = conn.execute('SELECT * FROM items WHERE barcode = ?', (barcode,)).fetchone()
    conn.close()
    
    if item:
        return jsonify({'success': True, 'item': dict(item)})
    else:
        return jsonify({'success': False, 'message': 'Item not found'}), 404

@app.route('/api/packing-lists', methods=['GET', 'POST'])
@login_required
def manage_packing_lists():
    """Get all packing lists or create new one"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        if not has_permission(current_user, 'create_packing_list'):
            return jsonify({'success': False, 'message': 'Permission denied'}), 403
        
        data = request.json
        list_name = data.get('list_name')
        
        cursor = conn.execute(
            'INSERT INTO packing_lists (list_name, created_by) VALUES (?, ?)',
            (list_name, current_user.id)
        )
        conn.commit()
        list_id = cursor.lastrowid
        conn.close()
        return jsonify({'success': True, 'list_id': list_id})
    
    else:  # GET
        if has_permission(current_user, 'view_all'):
            lists = conn.execute('''
                SELECT pl.*, u.username as created_by_name 
                FROM packing_lists pl
                LEFT JOIN users u ON pl.created_by = u.id
                ORDER BY pl.created_at DESC
            ''').fetchall()
        else:
            lists = conn.execute('''
                SELECT pl.*, u.username as created_by_name 
                FROM packing_lists pl
                LEFT JOIN users u ON pl.created_by = u.id
                WHERE pl.created_by = ?
                ORDER BY pl.created_at DESC
            ''', (current_user.id,)).fetchall()
        conn.close()
        return jsonify([dict(pl) for pl in lists])


# ============== THIRD PARTIES ROUTES ==============

@app.route('/api/third-parties', methods=['GET'])
@login_required
def get_third_parties():
    """Get all third parties"""
    try:
        conn = get_db_connection()
        third_parties = conn.execute('''
            SELECT third_party_id, name, type, city, address, 
                   contact_person, email, phone, created_at, updated_at 
            FROM third_parties 
            ORDER BY name ASC
        ''').fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'third_parties': [dict(row) for row in third_parties]
        })
    except Exception as e:
        print(f"❌ Error fetching third parties: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/third-parties', methods=['POST'])
@login_required
def add_third_party():
    """Add a new third party"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        party_type = data.get('type', '').strip()
        city = data.get('city', '').strip()
        address = data.get('address', '').strip()
        contact_person = data.get('contact_person', '').strip()
        email = data.get('email', '').strip()
        phone = data.get('phone', '').strip()
        
        if not name or not party_type:
            return jsonify({'success': False, 'message': 'Name and Type are required'}), 400
        
        conn = get_db_connection()
        cursor = conn.execute('''
            INSERT INTO third_parties (name, type, city, address, contact_person, email, phone) 
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (name, party_type, city, address, contact_person, email, phone))
        conn.commit()
        
        third_party_id = cursor.lastrowid
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Third party added successfully',
            'third_party_id': third_party_id
        })
    except Exception as e:
        print(f"❌ Error adding third party: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/third-parties/<int:third_party_id>', methods=['PUT'])
@login_required
def update_third_party(third_party_id):
    """Update a third party"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        party_type = data.get('type', '').strip()
        city = data.get('city', '').strip()
        address = data.get('address', '').strip()
        contact_person = data.get('contact_person', '').strip()
        email = data.get('email', '').strip()
        phone = data.get('phone', '').strip()
        
        if not name or not party_type:
            return jsonify({'success': False, 'message': 'Name and Type are required'}), 400
        
        conn = get_db_connection()
        conn.execute('''
            UPDATE third_parties 
            SET name = ?, type = ?, city = ?, address = ?, 
                contact_person = ?, email = ?, phone = ?, 
                updated_at = CURRENT_TIMESTAMP 
            WHERE third_party_id = ?
        ''', (name, party_type, city, address, contact_person, email, phone, third_party_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Third party updated successfully'
        })
    except Exception as e:
        print(f"❌ Error updating third party: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/third-parties/<int:third_party_id>', methods=['DELETE'])
@login_required
def delete_third_party(third_party_id):
    """Delete a third party"""
    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM third_parties WHERE third_party_id = ?', (third_party_id,))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Third party deleted successfully'
        })
    except Exception as e:
        print(f"❌ Error deleting third party: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500        

# ============== MISSION DETAILS ROUTES ==============

@app.route('/api/mission-details', methods=['GET'])
@login_required
def get_mission_details():
    """Get active mission details"""
    conn = get_db_connection()
    mission = conn.execute('''
        SELECT md.*, u.username as created_by_name
        FROM mission_details md
        LEFT JOIN users u ON md.created_by = u.id
        WHERE md.is_active = 1
        ORDER BY md.created_at DESC
        LIMIT 1
    ''').fetchone()
    conn.close()
    
    if mission:
        return jsonify({'success': True, 'data': dict(mission), 'exists': True})
    else:
        return jsonify({'success': True, 'data': None, 'exists': False})

@app.route('/api/mission-details/check', methods=['GET'])
@login_required
def check_mission_details():
    """Check if mission details exist"""
    conn = get_db_connection()
    count = conn.execute('SELECT COUNT(*) as count FROM mission_details WHERE is_active = 1').fetchone()['count']
    conn.close()
    
    return jsonify({'exists': count > 0})

@app.route('/api/mission-details', methods=['POST'])
@login_required
def create_mission_details():
    """Create new mission details (any user can create if none exist)"""
    conn = get_db_connection()
    
    # Check if mission details already exist
    existing = conn.execute('SELECT COUNT(*) as count FROM mission_details WHERE is_active = 1').fetchone()['count']
    
    if existing > 0:
        # Only admin can create when one already exists
        if not has_permission(current_user, 'manage_all'):
            conn.close()
            return jsonify({'success': False, 'message': 'Only administrators can update mission details'}), 403
    
    data = request.json
    
    # Validate required fields
    required_fields = ['mission_name', 'mission_abbreviation', 'cover_period_months']
    for field in required_fields:
        if not data.get(field):
            conn.close()
            return jsonify({'success': False, 'message': f'{field} is required'}), 400
    
    try:
        # Deactivate any existing active mission
        conn.execute('UPDATE mission_details SET is_active = 0 WHERE is_active = 1')
        
        # Insert new mission details
        cursor = conn.execute('''
            INSERT INTO mission_details (
                mission_name, mission_abbreviation,
                lead_time_months, cover_period_months, security_stock_months,
                is_active, created_by
            ) VALUES (?, ?, ?, ?, ?, 1, ?)
        ''', (
            data.get('mission_name'),
            data.get('mission_abbreviation'),
            data.get('lead_time_months', 0),
            data.get('cover_period_months'),
            data.get('security_stock_months', 0),
            current_user.id
        ))
        
        conn.commit()
        mission_id = cursor.lastrowid
        conn.close()
        
        return jsonify({'success': True, 'message': 'Mission details saved successfully', 'id': mission_id})
    
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/mission-details/<int:mission_id>', methods=['PUT'])
@login_required
def update_mission_details(mission_id):
    """Update mission details (only admin/HQ)"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Only administrators can update mission details'}), 403
    
    conn = get_db_connection()
    data = request.json
    
    # Validate required fields
    required_fields = ['mission_name', 'mission_abbreviation', 'cover_period_months']
    for field in required_fields:
        if not data.get(field):
            conn.close()
            return jsonify({'success': False, 'message': f'{field} is required'}), 400
    
    try:
        conn.execute('''
            UPDATE mission_details SET
                mission_name = ?,
                mission_abbreviation = ?,
                lead_time_months = ?,
                cover_period_months = ?,
                security_stock_months = ?
            WHERE id = ?
        ''', (
            data.get('mission_name'),
            data.get('mission_abbreviation'),
            data.get('lead_time_months', 0),
            data.get('cover_period_months'),
            data.get('security_stock_months', 0),
            mission_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Mission details updated successfully'})
    
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============== PROJECTS ROUTES ==============

@app.route('/api/projects', methods=['GET'])
@login_required
def get_projects():
    """Get all active projects with mission abbreviation/name from mission_details (single active record)."""
    conn = get_db_connection()
    projects = conn.execute('''
        SELECT p.*, u.username as created_by_name,
               md.mission_abbreviation, md.mission_name
        FROM projects p
        LEFT JOIN users u ON p.created_by = u.id
        LEFT JOIN mission_details md ON md.is_active = 1
        WHERE p.is_active = 1
        ORDER BY p.display_order, p.created_at
    ''').fetchall()
    conn.close()
    return jsonify({'success': True, 'data': [dict(p) for p in projects]})

@app.route('/api/projects/<int:project_id>', methods=['GET'])
@login_required
def get_project(project_id):
    """Get single project"""
    conn = get_db_connection()
    project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
    conn.close()
    
    if project:
        return jsonify({'success': True, 'data': dict(project)})
    else:
        return jsonify({'success': False, 'message': 'Project not found'}), 404

@app.route('/api/projects', methods=['POST'])
@login_required
def create_project():
    """Create new project (unlimited)"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Only administrators can create projects'}), 403
    
    conn = get_db_connection()
    data = request.json
    
    # Validate required fields
    if not data.get('project_name') or not data.get('project_code'):
        conn.close()
        return jsonify({'success': False, 'message': 'Project name and code are required'}), 400
    
    try:
        # Get current max display order
        max_order = conn.execute('SELECT MAX(display_order) as max_order FROM projects WHERE is_active = 1').fetchone()['max_order']
        next_order = (max_order or 0) + 1
        
        cursor = conn.execute('''
            INSERT INTO projects (
                project_name, project_code, description,
                display_order, created_by
            ) VALUES (?, ?, ?, ?, ?)
        ''', (
            data.get('project_name'),
            data.get('project_code').upper(),
            data.get('description', ''),
            next_order,
            current_user.id
        ))
        
        conn.commit()
        project_id = cursor.lastrowid
        conn.close()
        
        return jsonify({'success': True, 'message': 'Project created successfully', 'id': project_id})
    
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'success': False, 'message': 'Project code already exists'}), 400
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/projects/<int:project_id>', methods=['PUT'])
@login_required
def update_project(project_id):
    """Update project"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Only administrators can update projects'}), 403
    
    conn = get_db_connection()
    data = request.json
    
    try:
        conn.execute('''
            UPDATE projects SET
                project_name = ?,
                project_code = ?,
                description = ?
            WHERE id = ?
        ''', (
            data.get('project_name'),
            data.get('project_code').upper(),
            data.get('description', ''),
            project_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Project updated successfully'})
    
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'success': False, 'message': 'Project code already exists'}), 400
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
@login_required
def delete_project(project_id):
    """Delete (deactivate) project"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Only administrators can delete projects'}), 403
    
    conn = get_db_connection()
    
    try:
        conn.execute('UPDATE projects SET is_active = 0 WHERE id = ?', (project_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Project deleted successfully'})
    
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/projects/reorder', methods=['POST'])
@login_required
def reorder_projects():
    """Reorder projects"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Only administrators can reorder projects'}), 403
    
    conn = get_db_connection()
    data = request.json
    project_ids = data.get('project_ids', [])
    
    try:
        for index, project_id in enumerate(project_ids, start=1):
            conn.execute('UPDATE projects SET display_order = ? WHERE id = ?', (index, project_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Projects reordered successfully'})
    
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============== USER MANAGEMENT ROUTES ==============

@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    """Get all users (admin only)"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    conn = get_db_connection()
    users = conn.execute('''
        SELECT id, username, role, language, created_at
        FROM users
        ORDER BY created_at DESC
    ''').fetchall()
    conn.close()
    
    return jsonify({
        'success': True, 
        'data': [dict(u) for u in users],
        'current_user_id': current_user.id
    })

@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    """Create new user (admin only)"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    role = data.get('role', '').strip()
    language = data.get('language', 'en')
    
    # Validate
    if not username or not password or not role:
        return jsonify({'success': False, 'message': 'Username, password, and role are required'}), 400
    
    if len(password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters'}), 400
    
    if role not in ['HQ', 'Coordinator', 'Manager', 'Supervisor']:
        return jsonify({'success': False, 'message': 'Invalid role'}), 400
    
    conn = get_db_connection()
    
    try:
        # Check if username exists
        existing = conn.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
        if existing:
            conn.close()
            return jsonify({'success': False, 'message': 'Username already exists'}), 400
        
        # Create user
        password_hash = generate_password_hash(password)
        conn.execute('''
            INSERT INTO users (username, password, role, language)
            VALUES (?, ?, ?, ?)
        ''', (username, password_hash, role, language))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'User created successfully'})
    
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
def update_user(user_id):
    """Update user (admin only)"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    data = request.json
    username = data.get('username', '').strip()
    role = data.get('role', '').strip()
    language = data.get('language', 'en')
    password = data.get('password', '').strip()
    
    # Validate
    if not username or not role:
        return jsonify({'success': False, 'message': 'Username and role are required'}), 400
    
    if role not in ['HQ', 'Coordinator', 'Manager', 'Supervisor']:
        return jsonify({'success': False, 'message': 'Invalid role'}), 400
    
    if password and len(password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters'}), 400
    
    conn = get_db_connection()
    
    try:
        # Check if username exists for another user
        existing = conn.execute('SELECT id FROM users WHERE username = ? AND id != ?', (username, user_id)).fetchone()
        if existing:
            conn.close()
            return jsonify({'success': False, 'message': 'Username already exists'}), 400
        
        # Update user
        if password:
            password_hash = generate_password_hash(password)
            conn.execute('''
                UPDATE users SET
                    username = ?,
                    password = ?,
                    role = ?,
                    language = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (username, password_hash, role, language, user_id))
        else:
            conn.execute('''
                UPDATE users SET
                    username = ?,
                    role = ?,
                    language = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (username, role, language, user_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'User updated successfully'})
    
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
def delete_user(user_id):
    """Delete user (admin only)"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    # Cannot delete self
    if user_id == current_user.id:
        return jsonify({'success': False, 'message': 'You cannot delete your own account'}), 400
    
    conn = get_db_connection()
    
    try:
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'User deleted successfully'})
    
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500
    
# ============== EXCEL EXPORT ROUTES ==============

@app.route('/api/export/packing-list/<int:list_id>')
@login_required
def export_packing_list(list_id):
    """Export packing list to Excel with custom formatting"""
    if not has_permission(current_user, 'export'):
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    conn = get_db_connection()
    
    # Get packing list info
    packing_list = conn.execute('''
        SELECT pl.*, u.username as created_by_name 
        FROM packing_lists pl
        LEFT JOIN users u ON pl.created_by = u.id
        WHERE pl.id = ?
    ''', (list_id,)).fetchone()
    
    if not packing_list:
        conn.close()
        return jsonify({'success': False, 'message': 'Packing list not found'}), 404
    
    # Get items in the packing list
    items = conn.execute('''
        SELECT i.barcode, i.name, i.location, pli.quantity
        FROM packing_list_items pli
        JOIN items i ON pli.item_id = i.id
        WHERE pli.packing_list_id = ?
    ''', (list_id,)).fetchall()
    conn.close()
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List"
    
    # Define colors from theme
    header_fill = PatternFill(start_color="1F3A8A", end_color="1F3A8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="1F3A8A")
    
    # Title
    ws['A1'] = f"Packing List: {packing_list['list_name']}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    # Metadata
    current_date = normalize_date(datetime.now())
    ws['A2'] = f"Date: {current_date}"
    ws['A3'] = f"Created by: {packing_list['created_by_name']}"
    ws['A4'] = f"Status: {packing_list['status']}"
    
    # Column headers
    headers = ['Barcode', 'Item Name', 'Location', 'Quantity']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row, item in enumerate(items, 7):
        ws.cell(row=row, column=1, value=item['barcode'])
        ws.cell(row=row, column=2, value=item['name'])
        ws.cell(row=row, column=3, value=item['location'])
        
        # Format quantity based on user's language/locale
        quantity_formatted = format_excel_number(item['quantity'], current_user.language)
        ws.cell(row=row, column=4, value=quantity_formatted)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    # Save file
    filename = f"packing_list_{list_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('data', filename)
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/api/basic-data/export', methods=['GET'])
@login_required
def export_basic_data():
    """Export basic_data to Excel"""
    if not has_permission(current_user, 'export'):
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    conn = get_db_connection()
    data = conn.execute('SELECT * FROM basic_data ORDER BY imported_at DESC').fetchall()
    conn.close()
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Basic Data"
    
    # Headers
    headers = [
        'Unique ID', 'Packing Ref', 'Line No', 'Item Code', 'Item Description',
        'Qty Unit Tot', 'Packaging', 'Parcel N°', 'Nb Parcels', 'Batch No',
        'Exp Date', 'Kg Total', 'dm3 Total', 'Transport Reception', 'Sub Folder',
        'Field Ref', 'Ref Op MSFL', 'Parcel Nb', 'Weight (kg)', 'Volume (m3)',
        'Invoice/Credit Note Ref', 'Estim Value (EU)', 'Imported At'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="1F3A8A", end_color="1F3A8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, record in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=record['unique_id'])
        ws.cell(row=row_idx, column=2, value=record['packing_ref'])
        ws.cell(row=row_idx, column=3, value=record['line_no'])
        ws.cell(row=row_idx, column=4, value=record['item_code'])
        ws.cell(row=row_idx, column=5, value=record['item_description'])
        ws.cell(row=row_idx, column=6, value=record['qty_unit_tot'])
        ws.cell(row=row_idx, column=7, value=record['packaging'])
        ws.cell(row=row_idx, column=8, value=record['parcel_no'])
        ws.cell(row=row_idx, column=9, value=record['nb_parcels'])
        ws.cell(row=row_idx, column=10, value=record['batch_no'])
        ws.cell(row=row_idx, column=11, value=record['exp_date'])
        ws.cell(row=row_idx, column=12, value=record['kg_total'])
        ws.cell(row=row_idx, column=13, value=record['dm3_total'])
        ws.cell(row=row_idx, column=14, value=record['transport_reception'])
        ws.cell(row=row_idx, column=15, value=record['sub_folder'])
        ws.cell(row=row_idx, column=16, value=record['field_ref'])
        ws.cell(row=row_idx, column=17, value=record['ref_op_msfl'])
        ws.cell(row=row_idx, column=18, value=record['parcel_nb'])
        ws.cell(row=row_idx, column=19, value=record['weight_kg'])
        ws.cell(row=row_idx, column=20, value=record['volume_m3'])
        ws.cell(row=row_idx, column=21, value=record['invoice_credit_note_ref'])
        ws.cell(row=row_idx, column=22, value=record['estim_value_eu'])
        ws.cell(row=row_idx, column=23, value=record['imported_at'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    filename = f"basic_data_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('data', filename)
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

# ============== EXCEL IMPORT ROUTES ==============

@app.route('/api/import/preview', methods=['POST'])
@login_required
def preview_import():
    """Preview Excel file before importing"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
    
    file = request.files['file']
    file_type = request.form.get('file_type', 'file1')
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            data_rows, columns = ExcelImporter.read_excel_file(filepath, file_type)
            preview = data_rows[:10]  # First 10 rows for preview
            
            return jsonify({
                'success': True,
                'preview': preview,
                'columns': columns,
                'total_rows': len(data_rows),
                'filename': filename
            })
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error reading file: {str(e)}'}), 500
    
    return jsonify({'success': False, 'message': 'Invalid file type. Only .xlsx and .xls allowed'}), 400

@app.route('/api/import/execute', methods=['POST'])
@login_required
def execute_import():
    """Execute the import of one or two Excel files"""
    if not has_permission(current_user, 'manage_items'):
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    data = request.json
    file1_name = data.get('file1')
    file2_name = data.get('file2')
    
    if not file1_name:
        return jsonify({'success': False, 'message': 'At least one file is required'}), 400
    
    try:
        # Read first file
        file1_path = os.path.join(app.config['UPLOAD_FOLDER'], file1_name)
        
        if not os.path.exists(file1_path):
            return jsonify({'success': False, 'message': 'File 1 not found'}), 404
        
        file1_data, _ = ExcelImporter.read_excel_file(file1_path, 'file1')
        
        # If second file provided, merge data
        if file2_name:
            file2_path = os.path.join(app.config['UPLOAD_FOLDER'], file2_name)
            
            if not os.path.exists(file2_path):
                return jsonify({'success': False, 'message': 'File 2 not found'}), 404
            
            file2_data, _ = ExcelImporter.read_excel_file(file2_path, 'file2')
            merged_data = ExcelImporter.merge_data(file1_data, file2_data)
        else:
            merged_data = file1_data
        
        # Import to database
        source_files = f"{file1_name}" + (f" + {file2_name}" if file2_name else "")
        imported_count, errors = ExcelImporter.import_to_database(
            merged_data, 
            source_files, 
            current_user.id
        )
        
        return jsonify({
            'success': True,
            'imported_count': imported_count,
            'total_rows': len(merged_data),
            'errors': errors[:10] if errors else []  # Return first 10 errors
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'Import failed: {str(e)}'}), 500

@app.route('/api/basic-data', methods=['GET'])
@login_required
def get_basic_data():
    """Get all basic_data records with pagination"""
    conn = get_db_connection()
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    search = request.args.get('search', '', type=str)
    offset = (page - 1) * per_page
    
    # Build search query
    where_clause = ""
    params = []
    
    if search:
        where_clause = """
            WHERE bd.packing_ref LIKE ? OR 
                  bd.item_code LIKE ? OR 
                  bd.item_description LIKE ?
        """
        search_term = f"%{search}%"
        params = [search_term, search_term, search_term]
    
    # Get total count
    count_query = f"SELECT COUNT(*) as count FROM basic_data bd {where_clause}"
    total = conn.execute(count_query, params).fetchone()['count']
    
    # Get paginated data
    data_query = f"""
        SELECT bd.*, u.username as imported_by_name
        FROM basic_data bd
        LEFT JOIN users u ON bd.imported_by = u.id
        {where_clause}
        ORDER BY bd.imported_at DESC
        LIMIT ? OFFSET ?
    """
    data = conn.execute(data_query, params + [per_page, offset]).fetchall()
    
    conn.close()
    
    return jsonify({
        'success': True,
        'data': [dict(row) for row in data],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page
    })

@app.route('/api/basic-data/<int:record_id>', methods=['GET'])
@login_required
def get_basic_data_by_id(record_id):
    """Get a single basic_data record by ID"""
    conn = get_db_connection()
    record = conn.execute('''
        SELECT bd.*, u.username as imported_by_name
        FROM basic_data bd
        LEFT JOIN users u ON bd.imported_by = u.id
        WHERE bd.id = ?
    ''', (record_id,)).fetchone()
    conn.close()
    
    if record:
        return jsonify({'success': True, 'data': dict(record)})
    else:
        return jsonify({'success': False, 'message': 'Record not found'}), 404



# ============== BACKUP & RESTORE ROUTES ==============

import zipfile
import hashlib
import json

@app.route('/api/backup/create', methods=['POST'])
@login_required
def create_backup():
    """Create a backup zip file"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Only administrators can create backups'}), 403
    
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Get mission abbreviation for filename
        conn = get_db_connection()
        mission = conn.execute('SELECT mission_abbreviation FROM mission_details WHERE is_active = 1 LIMIT 1').fetchone()
        conn.close()
        
        mission_code = mission['mission_abbreviation'] if mission else 'MIDFLOW'
        backup_filename = f"midflow_backup_{mission_code}_{timestamp}.zip"
        backup_path = os.path.join('data', 'backups', backup_filename)
        
        # Create backups directory
        os.makedirs('data/backups', exist_ok=True)
        
        # Check if database exists
        if not os.path.exists(DATABASE):
            return jsonify({'success': False, 'message': 'Database file not found'}), 404
        
        # Compute SHA-256 hash
        sha256_hash = hashlib.sha256()
        db_size = os.path.getsize(DATABASE)
        with open(DATABASE, 'rb') as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        db_sha256 = sha256_hash.hexdigest()
        
        # Create backup metadata
        backup_meta = {
            "app_name": "MidFlow",
            "app_version": "1.0.0",
            "backup_date": datetime.now().isoformat(),
            "mission_code": mission_code,
            "created_by": current_user.username,
            "db_file": "inventory.db",
            "db_sha256": db_sha256,
            "db_size_bytes": db_size
        }
        
        # Create zip file
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add database
            zipf.write(DATABASE, 'inventory.db')
            
            # Add metadata
            metadata_json = json.dumps(backup_meta, indent=2)
            zipf.writestr("backup_meta.json", metadata_json)
        
        return jsonify({
            'success': True,
            'message': 'Backup created successfully',
            'filename': backup_filename,
            'size': os.path.getsize(backup_path),
            'download_url': f'/api/backup/download/{backup_filename}'
        })
    
    except Exception as e:
        print(f"❌ Error creating backup: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/backup/download/<filename>')
@login_required
def download_backup(filename):
    """Download a backup file"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        backup_path = os.path.join('data', 'backups', filename)
        
        if not os.path.exists(backup_path):
            return jsonify({'success': False, 'message': 'Backup file not found'}), 404
        
        return send_file(
            backup_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/zip'
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/backup/list', methods=['GET'])
@login_required
def list_backups():
    """List all available backups"""
    try:
        backups_dir = os.path.join('data', 'backups')
        
        # Create directory if it doesn't exist
        if not os.path.exists(backups_dir):
            print(f'📁 Creating backups directory: {backups_dir}')
            os.makedirs(backups_dir, exist_ok=True)
            return jsonify({'success': True, 'backups': []})
        
        backups = []
        
        # List all zip files
        for filename in os.listdir(backups_dir):
            if filename.endswith('.zip'):
                filepath = os.path.join(backups_dir, filename)
                
                try:
                    file_stat = os.stat(filepath)
                    
                    backups.append({
                        'filename': filename,
                        'size': file_stat.st_size,
                        'created_at': datetime.fromtimestamp(file_stat.st_mtime).isoformat()
                    })
                except Exception as e:
                    print(f'⚠️ Error reading file {filename}: {e}')
                    continue
        
        # Sort by creation date (newest first)
        backups.sort(key=lambda x: x['created_at'], reverse=True)
        
        print(f'✅ Found {len(backups)} backup files in {backups_dir}')
        
        return jsonify({'success': True, 'backups': backups})
    
    except Exception as e:
        print(f'❌ Error listing backups: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/backup/restore', methods=['POST'])
@login_required
def restore_backup():
    """Restore from uploaded backup file"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Only administrators can restore backups'}), 403
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.endswith('.zip'):
            return jsonify({'success': False, 'message': 'Invalid file type. Only .zip files allowed'}), 400
        
        # Save uploaded file temporarily
        temp_zip = os.path.join('data', 'temp_restore.zip')
        file.save(temp_zip)
        
        # Extract and validate
        temp_dir = 'data/temp_restore'
        os.makedirs(temp_dir, exist_ok=True)
        
        with zipfile.ZipFile(temp_zip, 'r') as zip_ref:
            zip_contents = zip_ref.namelist()
            
            # Check if database exists in backup
            if 'inventory.db' not in zip_contents:
                os.remove(temp_zip)
                shutil.rmtree(temp_dir, ignore_errors=True)
                return jsonify({'success': False, 'message': 'Invalid backup file: database not found'}), 400
            
            # Load metadata if available
            backup_meta = None
            if 'backup_meta.json' in zip_contents:
                metadata_content = zip_ref.read('backup_meta.json')
                backup_meta = json.loads(metadata_content)
            
            # Extract all files
            zip_ref.extractall(temp_dir)
        
        # Verify database file
        extracted_db = os.path.join(temp_dir, 'inventory.db')
        
        if not os.path.exists(extracted_db):
            os.remove(temp_zip)
            shutil.rmtree(temp_dir, ignore_errors=True)
            return jsonify({'success': False, 'message': 'Database file not found in backup'}), 400
        
        # Verify SHA-256 hash if metadata exists
        if backup_meta and 'db_sha256' in backup_meta:
            sha256_hash = hashlib.sha256()
            with open(extracted_db, 'rb') as f:
                for byte_block in iter(lambda: f.read(4096), b""):
                    sha256_hash.update(byte_block)
            computed_hash = sha256_hash.hexdigest()
            expected_hash = backup_meta['db_sha256']
            
            if computed_hash != expected_hash:
                os.remove(temp_zip)
                shutil.rmtree(temp_dir, ignore_errors=True)
                return jsonify({
                    'success': False,
                    'message': 'Database integrity check failed! File may be corrupted.',
                    'expected_hash': expected_hash,
                    'computed_hash': computed_hash
                }), 400
        
        # Create safety backup of current database
        if os.path.exists(DATABASE):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safety_backup = os.path.join('data', 'backups', f'pre_restore_safety_{timestamp}.db')
            shutil.copy(DATABASE, safety_backup)
        
        # Replace current database
        shutil.copy(extracted_db, DATABASE)
        
        # Clean up
        os.remove(temp_zip)
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        # Count records in key tables
        conn = get_db_connection()
        tables_info = []
        
        for table in ['users', 'projects', 'mission_details', 'items']:
            try:
                count = conn.execute(f'SELECT COUNT(*) as count FROM {table}').fetchone()['count']
                tables_info.append({'table': table, 'count': count})
            except:
                pass
        
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Database restored successfully',
            'backup_info': backup_meta,
            'tables': tables_info
        })
    
    except Exception as e:
        print(f"❌ Error restoring backup: {e}")
        # Clean up on error
        if os.path.exists('data/temp_restore.zip'):
            os.remove('data/temp_restore.zip')
        if os.path.exists('data/temp_restore'):
            shutil.rmtree('data/temp_restore', ignore_errors=True)
        
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/backup/delete/<filename>', methods=['DELETE'])
@login_required
def delete_backup(filename):
    """Delete a backup file"""
    if not has_permission(current_user, 'manage_all'):
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        backup_path = os.path.join('data', 'backups', filename)
        
        if not os.path.exists(backup_path):
            return jsonify({'success': False, 'message': 'Backup file not found'}), 404
        
        os.remove(backup_path)
        
        return jsonify({'success': True, 'message': 'Backup deleted successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============== END USERS ROUTES ==============

@app.route('/api/end-users', methods=['GET'])
@login_required
def get_end_users():
    """Get all end users"""
    try:
        conn = get_db_connection()
        end_users = conn.execute('''
            SELECT end_user_id, name, user_type, created_at, updated_at 
            FROM end_users 
            ORDER BY name ASC
        ''').fetchall()
        conn.close()
        
        return jsonify({
            'success': True,
            'end_users': [dict(row) for row in end_users]
        })
    except Exception as e:
        print(f"❌ Error fetching end users: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/end-users', methods=['POST'])
@login_required
def add_end_user():
    """Add a new end user"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        user_type = data.get('user_type', '').strip()

        if not name:
            return jsonify({'success': False, 'message': 'Name is required'}), 400

        conn = get_db_connection()
        cursor = conn.execute('''
            INSERT INTO end_users (name, user_type)
            VALUES (?, ?)
        ''', (name, user_type))
        conn.commit()
        
        end_user_id = cursor.lastrowid
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'End user added successfully',
            'end_user_id': end_user_id
        })
    except Exception as e:
        print(f"❌ Error adding end user: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/end-users/<int:end_user_id>', methods=['PUT'])
@login_required
def update_end_user(end_user_id):
    """Update an end user"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        user_type = data.get('user_type', '').strip()

        if not name:
            return jsonify({'success': False, 'message': 'Name is required'}), 400

        conn = get_db_connection()
        conn.execute('''
            UPDATE end_users
            SET name = ?, user_type = ?, updated_at = CURRENT_TIMESTAMP
            WHERE end_user_id = ?
        ''', (name, user_type, end_user_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'End user updated successfully'
        })
    except Exception as e:
        print(f"❌ Error updating end user: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/end-users/<int:end_user_id>', methods=['DELETE'])
@login_required
def delete_end_user(end_user_id):
    """Delete an end user"""
    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM end_users WHERE end_user_id = ?', (end_user_id,))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'End user deleted successfully'
        })
    except Exception as e:
        print(f"❌ Error deleting end user: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500   


# ================================================================
#  ADD THESE ROUTES TO app.py
# ================================================================

from flask import jsonify, request
from flask_login import login_required
import sqlite3
from datetime import datetime

# ── Helper ────────────────────────────────────────────────────
# get_db_connection() is imported from database.py (absolute path, no override needed)

def _og_db():
    """Get DB connection with WAL mode and row factory."""
    conn = get_db_connection()                          
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=10000")
    except Exception:
        pass
    return conn

    
# ── Helper: insert lines (no total_price — it's VIRTUAL) ──────

_VALID_STATUSES = {None, 'Requested', 'Approved', 'Rejected', 'Shared with Supply', 'Waiting for Quotation'}
def _og_safe_status(val):
    """Return val if it passes the DB CHECK, else None."""
    if val in _VALID_STATUSES:
        return val
    return None

def _og_insert_lines(conn, order_id, order_number, lines, currency='EUR'):
    for i, line in enumerate(lines, start=1):
        conn.execute("""
            INSERT INTO order_lines (
                order_id, order_number, line_no,
                item_code, item_description,
                quantity, project, order_family,
                packaging, price_per_pack,
                validation_status, remarks,
                order_description, order_generation_date, requested_delivery_date,
                currency
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            order_id,
            order_number,
            i,
            line.get('item_code'),
            line.get('item_description'),
            line.get('quantity', 0),
            line.get('project'),
            line.get('order_family'),
            line.get('packaging'),
            line.get('price_per_pack', 0),
            _og_safe_status(line.get('validation_status')),
            line.get('remarks'),
            line.get('order_description'),
            line.get('order_generation_date'),
            line.get('requested_delivery_date'),
            currency,
        ))






# ── GET all orders ─────────────────────────────────────────────
@app.route('/api/orders', methods=['GET'])
@login_required
def api_get_orders():
    conn = None
    try:
        conn = _og_db()
        # orders PK is "id"
        orders = conn.execute(
            "SELECT * FROM orders ORDER BY created_at DESC"
        ).fetchall()
        result = []
        for o in orders:
            od = dict(o)
            lines = conn.execute(
                "SELECT * FROM order_lines WHERE order_id=? ORDER BY line_no",
                (o['id'],)
            ).fetchall()
            od['order_id'] = o['id']
            od['lines'] = [dict(l) for l in lines]
            # Currency: read from first line (order_lines is source of truth)
            if lines:
                od['currency'] = lines[0]['currency'] if 'currency' in lines[0].keys() else (od.get('currency', 'EUR'))
            else:
                od['currency'] = od.get('currency', 'EUR')
            result.append(od)
        return jsonify({'success': True, 'orders': result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET single order ───────────────────────────────────────────
@app.route('/api/orders/<int:order_id>', methods=['GET'])
@login_required
def api_get_order(order_id):
    conn = None
    try:
        conn = _og_db()
        o = conn.execute(
            "SELECT * FROM orders WHERE id=?", (order_id,)
        ).fetchone()
        if not o:
            return jsonify({'success': False, 'message': 'Not found'}), 404
        od = dict(o)
        od['order_id'] = o['id']
        lines = conn.execute(
            "SELECT * FROM order_lines WHERE order_id=? ORDER BY line_no",
            (order_id,)
        ).fetchall()
        od['lines'] = [dict(l) for l in lines]
        # Currency from first line
        if lines:
            od['currency'] = lines[0]['currency'] if 'currency' in lines[0].keys() else (od.get('currency', 'EUR'))
        else:
            od['currency'] = od.get('currency', 'EUR')
        return jsonify({'success': True, 'order': od})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Next auto-number ───────────────────────────────────────────
@app.route('/api/orders/next-number', methods=['GET'])
@login_required
def api_next_order_number():
    conn = None
    try:
        order_type = request.args.get('type', 'Local')
        project    = request.args.get('project', 'XXX')
        family     = request.args.get('family', 'Med')
        yy         = datetime.today().strftime('%y')
        if order_type != 'Local':
            return jsonify({'success': True, 'order_number': ''})
        conn = _og_db()
        rows = conn.execute(
            "SELECT order_number FROM orders WHERE order_type=? AND order_number LIKE ?",
            ('Local', f"{yy}/{project}/{family}/%")
        ).fetchall()
        number = f"{yy}/{project}/{family}/LP{len(rows)+1:02d}"
        return jsonify({'success': True, 'order_number': number})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── CREATE order ───────────────────────────────────────────────
@app.route('/api/orders', methods=['POST'])
@login_required
def api_create_order():
    conn = None
    try:
        data  = request.get_json()
        lines = data.pop('lines', [])
        conn  = _og_db()
        cur   = conn.execute("""
            INSERT INTO orders (
                order_number, order_type, order_description,
                order_family, order_project,
                stock_date, requested_delivery_date,
                order_generation_date, created_by, currency
            ) VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            data.get('order_number'),
            data.get('order_type'),
            data.get('order_description'),
            data.get('order_family'),
            data.get('order_project'),
            data.get('stock_date'),
            data.get('requested_delivery_date'),
            datetime.today().strftime('%Y-%m-%d'),
            getattr(current_user, 'id', None),
            data.get('currency', 'EUR')
        ))
        new_id = cur.lastrowid
        _og_insert_lines(conn, new_id, data.get('order_number'), lines, data.get('currency', 'EUR'))
        conn.commit()
        return jsonify({'success': True, 'order_id': new_id})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── UPDATE order ───────────────────────────────────────────────
@app.route('/api/orders/<int:order_id>', methods=['PUT'])
@login_required
def api_update_order(order_id):
    conn = None
    try:
        data  = request.get_json()
        lines = data.pop('lines', [])
        conn  = _og_db()
        conn.execute("""
            UPDATE orders SET
                order_number=?, order_type=?, order_description=?,
                order_family=?, order_project=?,
                stock_date=?, requested_delivery_date=?,
                currency=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (
            data.get('order_number'),
            data.get('order_type'),
            data.get('order_description'),
            data.get('order_family'),
            data.get('order_project'),
            data.get('stock_date'),
            data.get('requested_delivery_date'),
            data.get('currency', 'EUR'),
            order_id
        ))
        conn.execute("DELETE FROM order_lines WHERE order_id=?", (order_id,))
        _og_insert_lines(conn, order_id, data.get('order_number'), lines, data.get('currency', 'EUR'))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── DELETE order ───────────────────────────────────────────────
@app.route('/api/orders/<int:order_id>', methods=['DELETE'])
@login_required
def api_delete_order(order_id):
    conn = None
    try:
        conn = _og_db()
        # order_lines has ON DELETE CASCADE, but we delete explicitly to be safe
        conn.execute("DELETE FROM order_lines WHERE order_id=?", (order_id,))
        conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


         

# ══════════════════════════════════════════════════════════════
# CARGO RECEPTION API ROUTES  (v2 — full implementation)
# ══════════════════════════════════════════════════════════════

def _cr_db():
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=10000")
        conn.execute("PRAGMA foreign_keys=ON")
    except Exception:
        pass
    return conn


def _cr_next_reception_number(conn, mission_abbrev='MSF'):
    """Generate next reception number: YY/ABBREV/SR{seq}"""
    year = __import__('datetime').date.today().strftime('%y')
    prefix = f"{year}/{mission_abbrev}/SR"
    row = conn.execute(
        "SELECT COUNT(*) FROM stock_transactions WHERE reception_number LIKE ?",
        (prefix + '%',)
    ).fetchone()
    seq = (row[0] if row else 0) + 1
    # Also count distinct reception numbers to avoid gaps from same-parcel multi-item
    row2 = conn.execute(
        "SELECT COUNT(DISTINCT reception_number) FROM stock_transactions WHERE reception_number LIKE ?",
        (prefix + '%',)
    ).fetchone()
    seq = (row2[0] if row2 else 0) + 1
    return f"{prefix}{seq:04d}"


def _cr_mission_abbrev(conn):
    """Get first active mission abbreviation."""
    row = conn.execute(
        "SELECT mission_abbreviation FROM mission_details WHERE is_active=1 ORDER BY id LIMIT 1"
    ).fetchone()
    return row['mission_abbreviation'] if row else 'MSN'


def _cr_extract_project_code(conn, field_ref):
    """Find which project_code from projects table appears inside field_ref string.
    Example: field_ref='25/CH/CD502/PO06146' and project_code='CD502' → returns 'CD502'."""
    if not field_ref:
        return None
    try:
        codes = [r[0] for r in conn.execute(
            "SELECT project_code FROM projects WHERE is_active=1 AND project_code IS NOT NULL AND project_code != ''"
        ).fetchall()]
        field_str = str(field_ref)
        for code in codes:
            if code and code in field_str:
                return code
    except Exception:
        pass
    return None


# ── GET mission info for reception number ─────────────────────
@app.route('/api/cargo/mission-info', methods=['GET'])
@login_required
def cr_mission_info():
    conn = None
    try:
        conn = _cr_db()
        abbrev = _cr_mission_abbrev(conn)
        next_num = _cr_next_reception_number(conn, abbrev)
        return jsonify({'success': True, 'abbrev': abbrev, 'next_reception_number': next_num})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET all cargo summary records ─────────────────────────────
@app.route('/api/cargo/summary', methods=['GET'])
@login_required
def cr_get_summary():
    conn = None
    try:
        conn = _cr_db()
        order_type = request.args.get('order_type', '')
        session_id = request.args.get('session_id', '')
        query = "SELECT * FROM cargo_summary WHERE 1=1"
        params = []
        if order_type:
            query += " AND order_type=?"
            params.append(order_type)
        if session_id:
            query += " AND cargo_session_id=?"
            params.append(session_id)
        query += " ORDER BY id DESC"
        rows = conn.execute(query, params).fetchall()
        return jsonify({'success': True, 'records': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── POST save cargo summary (batch from Excel parse) ──────────
@app.route('/api/cargo/summary', methods=['POST'])
@login_required
def cr_save_summary():
    conn = None
    try:
        data = request.json or {}
        records = data.get('records', [])
        order_type = data.get('order_type', 'Internal')
        session_id = data.get('session_id', '')
        if not records:
            return jsonify({'success': False, 'message': 'No records provided'}), 400

        conn = _cr_db()
        inserted = 0
        for rec in records:
            # Build parcel_number: goods_reception + parcel_nb  OR  just parcel_nb
            gr = str(rec.get('goods_reception') or '').strip()
            pn = str(rec.get('parcel_nb') or '').strip()
            auto_parcel = (gr + pn) if gr else pn

            conn.execute('''
                INSERT OR REPLACE INTO cargo_summary
                (parcel_number, transport_reception, sub_folder, field_ref,
                 ref_op_msfl, goods_reception, parcel_nb,
                 weight_kg, volume_m3, invoice_credit_note_ref, estim_value_eu,
                 reception_status, order_type, notes, cargo_session_id, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
            ''', (
                auto_parcel or None,
                rec.get('transport_reception'),
                rec.get('sub_folder'),
                rec.get('field_ref'),
                rec.get('ref_op_msfl'),
                rec.get('goods_reception'),
                rec.get('parcel_nb'),
                rec.get('weight_kg'),
                rec.get('volume_m3'),
                rec.get('invoice_credit_note_ref'),
                rec.get('estim_value_eu'),
                'Pending',
                order_type,
                rec.get('notes', ''),
                session_id,
            ))
            inserted += 1
        conn.commit()
        return jsonify({'success': True, 'inserted': inserted})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── DELETE cargo summary session ──────────────────────────────
@app.route('/api/cargo/summary/session/<session_id>', methods=['DELETE'])
@login_required
def cr_delete_summary_session(session_id):
    conn = None
    try:
        conn = _cr_db()
        conn.execute("DELETE FROM cargo_summary WHERE cargo_session_id=?", (session_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET all packing list records ──────────────────────────────
@app.route('/api/cargo/packing-list', methods=['GET'])
@login_required
def cr_get_packing_list():
    conn = None
    try:
        conn = _cr_db()
        session_id = request.args.get('session_id', '')
        packing_ref = request.args.get('packing_ref', '')
        query = "SELECT * FROM cargo_packing_list WHERE 1=1"
        params = []
        if session_id:
            query += " AND cargo_session_id=?"
            params.append(session_id)
        if packing_ref:
            query += " AND packing_ref=?"
            params.append(packing_ref)
        query += " ORDER BY packing_ref, parcel_nb, line_no"
        rows = conn.execute(query, params).fetchall()
        return jsonify({'success': True, 'records': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── POST save packing list → basic_data (direct merge, no JOIN staging) ──
@app.route('/api/cargo/packing-list', methods=['POST'])
@login_required
def cr_save_packing_list():
    conn = None
    try:
        data      = request.json or {}
        records   = data.get('records', [])
        session_id= data.get('session_id', '')
        order_type= data.get('order_type', 'International')
        if not records:
            return jsonify({'success': False, 'message': 'No records provided'}), 400

        conn = _cr_db()

        # ── Build cargo_summary lookup ─────────────────────────────────────
        # Primary key: (str(goods_reception), str(parcel_nb))  → exact match
        # Fallback key: str(goods_reception)                   → any parcel for that ref
        cs_all      = conn.execute('SELECT * FROM cargo_summary').fetchall()
        cs_by_exact = {}   # (goods_reception_str, parcel_nb_str) → row dict
        cs_by_ref   = {}   # goods_reception_str                   → row dict

        for row in cs_all:
            rd  = dict(row)
            gr  = str(rd.get('goods_reception', '') or '').strip()
            pnb = str(rd.get('parcel_nb', '') or '').strip()
            if gr:
                cs_by_exact[(gr, pnb)] = rd
                cs_by_ref[gr]          = rd   # last row for this ref as fallback

        # Preserve already-received parcels: do not reset status when re-uploading
        received_pn = set()
        for r in conn.execute(
            "SELECT DISTINCT parcel_number FROM basic_data WHERE reception_number IS NOT NULL AND parcel_number IS NOT NULL"
        ).fetchall():
            if r['parcel_number']:
                received_pn.add(str(r['parcel_number']))

        bd_inserted = 0

        for rec in records:
            packing_ref = str(rec.get('packing_ref', '') or '').strip()
            line_no     = rec.get('line_no')
            parcel_nb   = rec.get('parcel_nb')

            # Barcode = packing_ref + parcel_nb  (matches physical label)
            auto_pn   = (packing_ref + str(parcel_nb or '')).strip() or None

            # unique_id includes line_no so multiple items per parcel are kept
            unique_id = f"{packing_ref}_{line_no}_{parcel_nb}"

            # Look up cargo_summary: try exact (ref, parcel_nb) first, then just ref
            cs = cs_by_exact.get((packing_ref, str(parcel_nb or '').strip()), {})
            if not cs:
                cs = cs_by_ref.get(packing_ref, {})

            # ── Insert into packing_list (staging table, Parcel_number PK) ─
            # NOTE: packing_list has PRIMARY KEY on Parcel_number so only the
            # last item per parcel survives here — that is intentional for the
            # manifest summary.  basic_data below holds ALL item rows.
            try:
                conn.execute('''
                    INSERT OR REPLACE INTO packing_list
                    (Parcel_number, Packing_ref, Line_no, Item_code, Item_description,
                     Qty_unit_tot, Packaging, Parcel_n, Nb_parcels, Batch_no,
                     Exp_date, Kg_total, Dm3_total, Parcel_nb, cargo_session_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ''', (
                    auto_pn, packing_ref, line_no,
                    rec.get('item_code'), rec.get('item_description'),
                    rec.get('qty_unit_tot'), rec.get('packaging'),
                    rec.get('parcel_n'), rec.get('nb_parcels'),
                    rec.get('batch_no'), rec.get('exp_date'),
                    rec.get('kg_total'), rec.get('dm3_total'),
                    parcel_nb, session_id,
                ))
            except Exception:
                pass  # packing_list is secondary; basic_data is the real store

            # ── Insert ALL rows into basic_data (unique per item line) ─────
            project_code = _cr_extract_project_code(conn, cs.get('field_ref'))
            conn.execute('''
                INSERT OR REPLACE INTO basic_data
                (unique_id, packing_ref, line_no, item_code, item_description,
                 qty_unit_tot, packaging, parcel_no, nb_parcels, batch_no,
                 exp_date, kg_total, dm3_total,
                 transport_reception, sub_folder, field_ref, ref_op_msfl,
                 parcel_nb, weight_kg, volume_m3,
                 invoice_credit_note_ref, estim_value_eu,
                 parcel_number, reception_status, order_type, cargo_session_id,
                 source_file, imported_by, project_code)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                unique_id,
                packing_ref or None,
                line_no,
                rec.get('item_code'),
                rec.get('item_description'),
                rec.get('qty_unit_tot'),
                rec.get('packaging'),
                rec.get('parcel_n'),
                rec.get('nb_parcels'),
                rec.get('batch_no'),
                rec.get('exp_date'),
                rec.get('kg_total'),
                rec.get('dm3_total'),
                cs.get('transport_reception'),
                cs.get('sub_folder'),
                cs.get('field_ref'),
                cs.get('ref_op_msfl'),
                str(parcel_nb or ''),
                cs.get('weight_kg'),
                cs.get('volume_m3'),
                cs.get('invoice_credit_note_ref'),
                cs.get('estim_value_eu'),
                auto_pn,          # parcel_number = barcode on physical label
                'Received' if auto_pn and auto_pn in received_pn else 'Pending',
                order_type,
                session_id,
                'Excel Import',
                current_user.id,
                project_code,
            ))
            bd_inserted += 1

        conn.commit()
        return jsonify({'success': True, 'pl_inserted': bd_inserted, 'bd_inserted': bd_inserted})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── DELETE packing list session (also clears basic_data for that session) ─
@app.route('/api/cargo/packing-list/session/<session_id>', methods=['DELETE'])
@login_required
def cr_delete_packing_session(session_id):
    conn = None
    try:
        conn = _cr_db()
        # Clear staging tables and basic_data rows that have NOT been received
        conn.execute("DELETE FROM packing_list WHERE cargo_session_id=?",   (session_id,))
        conn.execute("DELETE FROM basic_data   WHERE cargo_session_id=? AND (reception_status IS NULL OR reception_status='Pending')",
                     (session_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET basic_data records ────────────────────────────────────
@app.route('/api/cargo/basic-data', methods=['GET'])
@login_required
def cr_get_basic_data():
    conn = None
    try:
        conn = _cr_db()
        session_id  = request.args.get('session_id', '')
        order_type  = request.args.get('order_type', '')
        status      = request.args.get('status', '')
        q = "SELECT * FROM basic_data WHERE 1=1"
        params = []
        if session_id:
            q += " AND cargo_session_id=?"; params.append(session_id)
        if order_type:
            q += " AND order_type=?";       params.append(order_type)
        if status:
            q += " AND reception_status=?"; params.append(status)
        q += " ORDER BY packing_ref, CAST(parcel_nb AS INTEGER), line_no"
        rows = conn.execute(q, params).fetchall()
        return jsonify({'success': True, 'records': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET items for a specific parcel (from basic_data) ─────────
@app.route('/api/cargo/packing-list/<parcel_number>', methods=['GET'])
@login_required
def cr_get_packing_for_parcel(parcel_number):
    conn = None
    try:
        conn = _cr_db()
        rows = conn.execute(
            "SELECT * FROM basic_data WHERE parcel_number=? ORDER BY line_no",
            (parcel_number,)
        ).fetchall()
        return jsonify({'success': True, 'items': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET unique parcels (distinct parcel_number from basic_data) ─
@app.route('/api/cargo/parcels', methods=['GET'])
@login_required
def cr_get_parcels():
    conn = None
    try:
        conn = _cr_db()
        session_id = request.args.get('session_id', '')
        order_type = request.args.get('order_type', '')
        # NOTE: basic_data has 'Parcel_number' (capital P, legacy schema).
        # Explicit AS alias forces lowercase keys in dict/JSON so JS can read r.parcel_number.
        q = '''
            SELECT parcel_number  AS parcel_number,
                   field_ref, packing_ref, parcel_nb,
                   transport_reception, weight_kg, volume_m3, estim_value_eu,
                   CASE WHEN MAX(reception_number) IS NOT NULL OR MAX(reception_status) = 'Received'
                        THEN 'Received' ELSE 'Pending' END AS reception_status,
                   MAX(reception_number) AS reception_number,
                   MAX(received_at)      AS received_at,
                   MAX(pallet_number)    AS pallet_number,
                   MAX(cargo_session_id) AS cargo_session_id,
                   MAX(parcel_note)      AS parcel_note,
                   MAX(project_code)     AS project_code,
                   MAX(order_type)       AS order_type,
                   COUNT(*)              AS item_count
            FROM basic_data WHERE parcel_number IS NOT NULL AND parcel_number != ""
        '''
        params = []
        if session_id:
            q += " AND cargo_session_id=?"; params.append(session_id)
        if order_type:
            q += " AND order_type=?";       params.append(order_type)
        q += " GROUP BY parcel_number ORDER BY CAST(parcel_nb AS INTEGER)"
        rows = conn.execute(q, params).fetchall()
        return jsonify({'success': True, 'parcels': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── POST receive a parcel (button click or barcode scan) ───────
@app.route('/api/cargo/receive-parcel', methods=['POST'])
@login_required
def cr_receive_parcel_v2():
    conn = None
    try:
        data        = request.json or {}
        parcel_num  = str(data.get('parcel_number', '')).strip()
        pallet      = data.get('pallet_number', '')
        notes       = data.get('notes', '')
        order_type  = data.get('order_type', 'International')
        session_id  = data.get('session_id', '')
        exp_date    = data.get('exp_date', '')        # reception-time expiry date (or 'N/A')
        batch_no    = data.get('batch_no', '')        # reception-time batch number (optional)

        if not parcel_num:
            return jsonify({'success': False, 'message': 'parcel_number required'}), 400

        conn = _cr_db()

        # Look up parcel rows in basic_data — NO session filter: parcels from any session are receivable
        rows = conn.execute(
            "SELECT * FROM basic_data WHERE parcel_number=?", [parcel_num]
        ).fetchall()

        if not rows:
            return jsonify({'success': False, 'message': 'not_found', 'parcel_number': parcel_num}), 404

        # Check if already received
        if all(dict(r)['reception_status'] == 'Received' for r in rows):
            first = dict(rows[0])
            return jsonify({
                'success': False,
                'message': 'already_received',
                'reception_number': first.get('reception_number', ''),
                'received_at': first.get('received_at', ''),
                'parcel_number': parcel_num
            }), 409

        # Generate reception number
        abbrev     = _cr_mission_abbrev(conn)
        recep_num  = _cr_next_reception_number(conn, abbrev)

        # Create stock_transaction records (one per item line)
        for row in rows:
            rd = dict(row)
            # Use reception-time exp_date/batch_no if provided; fall back to packing-list values
            eff_exp_date = exp_date if exp_date else rd.get('exp_date', '')
            eff_batch_no = batch_no if batch_no else rd.get('batch_no', '')
            conn.execute('''
                INSERT INTO stock_transactions
                (reception_number, transaction_type, parcel_number,
                 packing_ref, line_no, item_code, item_description,
                 qty_received, packaging, batch_no, exp_date,
                 order_number, field_ref, pallet_number,
                 transport_reception, weight_kg, volume_m3, estim_value_eu,
                 mission_abbreviation, received_by, cargo_session_id, notes,
                 project_code)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                recep_num, 'RECEPTION', parcel_num,
                rd.get('packing_ref'), rd.get('line_no'),
                rd.get('item_code'), rd.get('item_description'),
                rd.get('qty_unit_tot'), rd.get('packaging'),
                eff_batch_no, eff_exp_date,
                rd.get('field_ref'), rd.get('field_ref'),
                pallet, rd.get('transport_reception'),
                rd.get('weight_kg'), rd.get('volume_m3'),
                rd.get('estim_value_eu'), abbrev,
                current_user.id, rd.get('cargo_session_id', session_id), notes,
                rd.get('project_code')
            ))

        # Update basic_data rows — mark received + record qty + exp/batch
        conn.execute('''
            UPDATE basic_data
            SET reception_status='Received',
                reception_number=?,
                received_at=CURRENT_TIMESTAMP,
                received_by=?,
                pallet_number=?,
                qty_received=qty_unit_tot,
                exp_date_received=CASE WHEN ? != '' THEN ? ELSE exp_date END,
                batch_no_received=CASE WHEN ? != '' THEN ? ELSE batch_no END
            WHERE parcel_number=?
        ''', (recep_num, current_user.id, pallet,
              exp_date, exp_date,
              batch_no, batch_no,
              parcel_num))

        # Also update cargo_summary if it has this parcel
        conn.execute('''
            UPDATE cargo_summary
            SET reception_status='Received',
                received_at=CURRENT_TIMESTAMP,
                received_by=?,
                notes=?
            WHERE parcel_number=?
        ''', (current_user.id, notes, parcel_num))

        conn.commit()

        first = dict(rows[0])
        return jsonify({
            'success': True,
            'reception_number': recep_num,
            'parcel_number': parcel_num,
            'field_ref': first.get('field_ref', ''),
            'item_count': len(rows),
            'pallet_number': pallet
        })
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── POST un-receive a parcel ──────────────────────────────────
@app.route('/api/cargo/unreceive-parcel', methods=['POST'])
@login_required
def cr_unreceive_parcel():
    conn = None
    try:
        data       = request.json or {}
        parcel_num = str(data.get('parcel_number', '')).strip()
        if not parcel_num:
            return jsonify({'success': False, 'message': 'parcel_number required'}), 400
        conn = _cr_db()
        # Get reception_number before deleting
        row = conn.execute(
            "SELECT reception_number FROM basic_data WHERE parcel_number=? LIMIT 1", (parcel_num,)
        ).fetchone()
        recep_num = row['reception_number'] if row else None

        conn.execute("DELETE FROM stock_transactions WHERE parcel_number=? AND reception_number=?",
                     (parcel_num, recep_num))
        conn.execute('''
            UPDATE basic_data
            SET reception_status='Pending', reception_number=NULL,
                received_at=NULL, received_by=NULL, pallet_number=NULL
            WHERE parcel_number=?
        ''', (parcel_num,))
        conn.execute('''
            UPDATE cargo_summary
            SET reception_status='Pending', received_at=NULL, received_by=NULL
            WHERE parcel_number=?
        ''', (parcel_num,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── PATCH update note for a parcel (pre-reception or any time) ─
@app.route('/api/cargo/parcel-note', methods=['PATCH'])
@login_required
def cr_update_parcel_note():
    conn = None
    try:
        data       = request.json or {}
        parcel_num = str(data.get('parcel_number', '')).strip()
        note       = data.get('note', '')
        if not parcel_num:
            return jsonify({'success': True})  # silently ignore stale onblur with empty parcel
        conn = _cr_db()
        conn.execute(
            "UPDATE basic_data SET parcel_note=? WHERE parcel_number=?",
            (note or None, parcel_num)
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── POST save local order parcels directly to basic_data ──────
@app.route('/api/cargo/local-parcels', methods=['POST'])
@login_required
def cr_save_local_parcels():
    conn = None
    try:
        data       = request.json or {}
        records    = data.get('records', [])
        session_id = data.get('session_id', '')
        if not records:
            return jsonify({'success': False, 'message': 'No records provided'}), 400

        conn = _cr_db()

        # Preserve already-received parcel_numbers
        received_pn = set()
        for r in conn.execute(
            "SELECT DISTINCT parcel_number FROM basic_data WHERE reception_number IS NOT NULL AND parcel_number IS NOT NULL"
        ).fetchall():
            if r['parcel_number']:
                received_pn.add(str(r['parcel_number']))

        inserted = 0
        for rec in records:
            parcel_num = str(rec.get('parcel_number', '') or '').strip()
            if not parcel_num:
                continue
            field_ref    = str(rec.get('field_ref', '') or '').strip() or None
            project_code = rec.get('project_code') or _cr_extract_project_code(conn, field_ref)
            status       = 'Received' if parcel_num in received_pn else 'Pending'
            # unique_id for local = LOCAL_{parcel_number} (one logical row per parcel)
            unique_id = f"LOCAL_{parcel_num}"

            conn.execute('''
                INSERT OR REPLACE INTO basic_data
                (unique_id, packing_ref, line_no, item_code,
                 transport_reception, field_ref, project_code,
                 parcel_nb, weight_kg, volume_m3,
                 invoice_credit_note_ref, estim_value_eu,
                 parcel_number, parcel_note, reception_status, order_type,
                 cargo_session_id, source_file, imported_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                unique_id, None, 1, None,
                rec.get('transport_reception'),
                field_ref,
                project_code,
                parcel_num,             # parcel_nb  = same as barcode for local
                rec.get('weight_kg'),
                rec.get('volume_m3'),
                rec.get('invoice_credit_note_ref'),
                rec.get('estim_value_eu'),
                parcel_num,             # parcel_number = barcode
                rec.get('notes') or None,
                status,
                'Local',
                session_id,
                'Manual Entry',
                current_user.id,
            ))
            inserted += 1

        conn.commit()
        return jsonify({'success': True, 'inserted': inserted})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── POST recalculate project codes for all existing basic_data rows ─
@app.route('/api/cargo/recalculate-projects', methods=['POST'])
@login_required
def cr_recalculate_projects():
    conn = None
    try:
        conn = _cr_db()
        rows = conn.execute(
            "SELECT DISTINCT parcel_number, field_ref FROM basic_data WHERE field_ref IS NOT NULL AND parcel_number IS NOT NULL"
        ).fetchall()
        updated = 0
        for row in rows:
            pcode = _cr_extract_project_code(conn, row['field_ref'])
            if pcode:
                conn.execute(
                    "UPDATE basic_data SET project_code=? WHERE parcel_number=? AND (project_code IS NULL OR project_code='')",
                    (pcode, row['parcel_number'])
                )
                updated += 1
        conn.commit()

        # Sync stock_transactions.project_code from basic_data for any rows that are still NULL
        conn.execute('''
            UPDATE stock_transactions
            SET project_code = (
                SELECT bd.project_code FROM basic_data bd
                WHERE bd.parcel_number = stock_transactions.parcel_number
                  AND bd.project_code IS NOT NULL AND bd.project_code != ''
                LIMIT 1
            )
            WHERE (project_code IS NULL OR project_code = '')
              AND parcel_number IS NOT NULL
        ''')
        conn.commit()

        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET local order lines for reception ───────────────────────
@app.route('/api/cargo/local-lines', methods=['GET'])
@login_required
def cr_get_local_lines():
    """Return all order_lines whose order is order_type='Local', with reception info."""
    conn = None
    try:
        conn = _cr_db()
        rows = conn.execute("""
            SELECT
                ol.line_id,
                ol.order_id,
                ol.order_number,
                ol.line_no,
                ol.item_code,
                ol.item_description,
                ol.quantity          AS qty_ordered,
                ol.packaging,
                ol.project           AS project_code,
                ol.remarks,
                ol.order_family,
                ol.currency,
                ol.price_per_pack,
                COALESCE(ol.qty_received, 0)                         AS qty_received,
                COALESCE(ol.quantity, 0) - COALESCE(ol.qty_received, 0) AS balance_qty,
                COALESCE(ol.reception_status, 'Pending')             AS reception_status,
                ol.exp_date_received,
                ol.batch_no_received,
                ol.received_at,
                COALESCE(ol.order_type,
                    (SELECT o.order_type FROM orders o
                     WHERE o.order_number = ol.order_number LIMIT 1)
                ) AS order_type
            FROM order_lines ol
            WHERE COALESCE(ol.order_type,
                    (SELECT o.order_type FROM orders o
                     WHERE o.order_number = ol.order_number LIMIT 1)
                  ) = 'Local'
            ORDER BY ol.order_number, ol.line_no
        """).fetchall()

        result = []
        for r in rows:
            d = dict(r)
            # Ensure numeric types
            d['qty_ordered']  = float(d['qty_ordered'])  if d['qty_ordered']  is not None else 0.0
            d['qty_received'] = float(d['qty_received']) if d['qty_received'] is not None else 0.0
            d['balance_qty']  = float(d['balance_qty'])  if d['balance_qty']  is not None else 0.0
            result.append(d)

        return jsonify({'success': True, 'lines': result, 'total': len(result)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── POST receive one order line (local reception) ─────────────
@app.route('/api/cargo/receive-line', methods=['POST'])
@login_required
def cr_receive_line():
    """Receive a specific quantity for a local order_line.
    Supports partial reception — can call multiple times until balance = 0.
    """
    conn = None
    try:
        data     = request.json or {}
        line_id  = data.get('line_id')
        qty      = data.get('qty_received')
        exp_date = str(data.get('exp_date', '') or '').strip()
        batch_no = str(data.get('batch_no', '') or '').strip()

        if not line_id:
            return jsonify({'success': False, 'message': 'line_id required'}), 400
        if qty is None:
            return jsonify({'success': False, 'message': 'qty_received required'}), 400
        if not exp_date:
            return jsonify({'success': False, 'message': 'exp_date required (enter date or N/A)'}), 400

        try:
            qty = float(qty)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'message': 'qty_received must be numeric'}), 400

        if qty <= 0:
            return jsonify({'success': False, 'message': 'qty_received must be > 0'}), 400

        conn = _cr_db()

        row = conn.execute(
            "SELECT * FROM order_lines WHERE line_id=?", [line_id]
        ).fetchone()

        if not row:
            return jsonify({'success': False, 'message': 'order line not found'}), 404

        rd = dict(row)
        qty_ordered  = float(rd.get('quantity') or 0)
        qty_prev     = float(rd.get('qty_received') or 0)
        new_qty_recv = qty_prev + qty

        if new_qty_recv > qty_ordered:
            return jsonify({
                'success': False,
                'message': f'Cannot receive {qty} — only {qty_ordered - qty_prev} remaining in balance'
            }), 400

        # Determine new status
        new_status = 'Fully Received' if new_qty_recv >= qty_ordered else 'Partial'

        # Generate reception number
        abbrev    = _cr_mission_abbrev(conn)
        recep_num = _cr_next_reception_number(conn, abbrev)

        # Create stock_transaction record
        # Note: order_lines column is 'project', not 'project_code'
        proj_code = rd.get('project') or rd.get('project_code') or None
        conn.execute('''
            INSERT INTO stock_transactions
            (reception_number, transaction_type, item_code, item_description,
             qty_received, packaging, batch_no, exp_date,
             order_number, field_ref,
             mission_abbreviation, received_by, notes, project_code)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            recep_num, 'RECEPTION',
            rd.get('item_code'), rd.get('item_description'),
            qty, rd.get('packaging'), batch_no, exp_date,
            rd.get('order_number'), rd.get('order_number'),
            abbrev, current_user.id, '', proj_code
        ))

        # Update order_lines
        conn.execute('''
            UPDATE order_lines
            SET qty_received     = ?,
                exp_date_received= ?,
                batch_no_received= ?,
                received_at      = CURRENT_TIMESTAMP,
                received_by      = ?,
                reception_status = ?
            WHERE line_id=?
        ''', (new_qty_recv, exp_date, batch_no or None,
              current_user.id, new_status, line_id))

        conn.commit()

        return jsonify({
            'success': True,
            'line_id': line_id,
            'reception_number': recep_num,
            'qty_ordered': qty_ordered,
            'qty_received': new_qty_recv,
            'balance_qty': qty_ordered - new_qty_recv,
            'reception_status': new_status,
            'exp_date_received': exp_date,
            'batch_no_received': batch_no
        })
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── DELETE a single cargo summary record ──────────────────────
@app.route('/api/cargo/summary/<int:summary_id>', methods=['DELETE'])
@login_required
def cr_delete_summary(summary_id):
    conn = None
    try:
        conn = _cr_db()
        conn.execute("DELETE FROM cargo_summary WHERE id=?", (summary_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET reception statistics ──────────────────────────────────
@app.route('/api/cargo/summary/stats', methods=['GET'])
@login_required
def cr_summary_stats():
    conn = None
    try:
        conn = _cr_db()
        session_id = request.args.get('session_id', '')
        base   = "FROM basic_data WHERE 1=1"
        params = []
        if session_id:
            base += " AND cargo_session_id=?"; params.append(session_id)

        # Count distinct parcels
        total   = conn.execute(f"SELECT COUNT(DISTINCT parcel_number) {base} AND parcel_number != ''", params).fetchone()[0]
        pending = conn.execute(f"SELECT COUNT(DISTINCT parcel_number) {base} AND reception_status='Pending' AND parcel_number != ''", params).fetchone()[0]
        done    = conn.execute(f"SELECT COUNT(DISTINCT parcel_number) {base} AND reception_status='Received' AND parcel_number != ''", params).fetchone()[0]

        # Weight totals
        w_total   = conn.execute(f"SELECT COALESCE(SUM(weight_kg),0) FROM cargo_summary WHERE 1=1{' AND cargo_session_id=?' if session_id else ''}", params).fetchone()[0]
        w_received= conn.execute(f"SELECT COALESCE(SUM(cs.weight_kg),0) FROM cargo_summary cs JOIN basic_data bd ON bd.parcel_number=cs.parcel_number WHERE bd.reception_status='Received'{' AND bd.cargo_session_id=?' if session_id else ''}", params).fetchone()[0]

        # Local order line stats (from order_lines table)
        lo_base = (
            "FROM order_lines ol LEFT JOIN orders o ON o.order_number = ol.order_number "
            "WHERE COALESCE(ol.order_type, o.order_type) = 'Local'"
        )
        lo_total   = conn.execute(f"SELECT COUNT(*) {lo_base}").fetchone()[0]
        lo_full    = conn.execute(f"SELECT COUNT(*) {lo_base} AND COALESCE(ol.reception_status,'Pending') = 'Fully Received'").fetchone()[0]
        lo_partial = conn.execute(f"SELECT COUNT(*) {lo_base} AND COALESCE(ol.reception_status,'Pending') = 'Partial'").fetchone()[0]

        return jsonify({
            'success': True,
            'total': total, 'pending': pending, 'received': done,
            'weight_total': round(float(w_total), 2),
            'weight_received': round(float(w_received), 2),
            'lo_total': lo_total, 'lo_full': lo_full, 'lo_partial': lo_partial,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET backorders calculation ────────────────────────────────
@app.route('/api/cargo/backorders', methods=['GET'])
@login_required
def cr_backorders():
    """Compare ordered qty (order_lines) vs received qty (stock_transactions)."""
    conn = None
    try:
        conn = _cr_db()
        rows = conn.execute('''
            SELECT
                ol.item_code,
                ol.item_description,
                o.order_number,
                SUM(ol.quantity) AS ordered_qty,
                COALESCE(st.received_qty, 0) AS received_qty,
                SUM(ol.quantity) - COALESCE(st.received_qty, 0) AS backorder_qty
            FROM order_lines ol
            JOIN orders o ON o.id = ol.order_id
            LEFT JOIN (
                SELECT item_code, order_number, SUM(qty_received) AS received_qty
                FROM stock_transactions
                WHERE transaction_type = 'RECEPTION'
                GROUP BY item_code, order_number
            ) st ON st.item_code = ol.item_code
               AND st.order_number = o.order_number
            GROUP BY ol.item_code, o.order_number
            HAVING backorder_qty > 0
            ORDER BY backorder_qty DESC
        ''').fetchall()
        return jsonify({'success': True, 'backorders': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET current stock summary ─────────────────────────────────
@app.route('/api/cargo/stock-summary', methods=['GET'])
@login_required
def cr_stock_summary():
    conn = None
    try:
        conn = _cr_db()
        rows = conn.execute('''
            SELECT item_code, item_description,
                   SUM(qty_received) AS total_received,
                   COUNT(DISTINCT reception_number) AS reception_count,
                   MAX(received_at) AS last_received
            FROM stock_transactions
            WHERE transaction_type = 'RECEPTION'
            GROUP BY item_code
            ORDER BY item_code
        ''').fetchall()
        return jsonify({'success': True, 'stock': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET reception history ─────────────────────────────────────
@app.route('/api/cargo/reception-history', methods=['GET'])
@login_required
def cr_reception_history():
    conn = None
    try:
        conn = _cr_db()
        rows = conn.execute('''
            SELECT st.*, u.username AS received_by_name
            FROM stock_transactions st
            LEFT JOIN users u ON u.id = st.received_by
            ORDER BY st.received_at DESC
            LIMIT 200
        ''').fetchall()
        return jsonify({'success': True, 'history': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── GET next pallet number suggestion ─────────────────────────
@app.route('/api/cargo/next-pallet', methods=['GET'])
@login_required
def cr_next_pallet():
    conn = None
    try:
        conn = _cr_db()
        row = conn.execute(
            "SELECT MAX(pallet_number) FROM stock_transactions WHERE pallet_number IS NOT NULL AND pallet_number != ''"
        ).fetchone()
        last = row[0] if row and row[0] else 'P000'
        # Extract number and increment
        import re
        m = re.search(r'\d+$', last)
        next_seq = (int(m.group()) + 1) if m else 1
        prefix   = re.sub(r'\d+$', '', last) if m else 'P'
        next_pallet = f"{prefix}{next_seq:03d}"
        return jsonify({'success': True, 'next_pallet': next_pallet})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════
#  MOVEMENTS — IN / OUT  (Phases 2 & 3)
# ═══════════════════════════════════════════════════════════════════════════

IN_TYPES = [
    ('IMSF',  'In MSF',                              None),
    ('ILP',   'In Local Purchase',                   None),
    ('IQRT',  'In from Quarantine',                  None),
    ('IDN',   'In Donation',                         'third_party'),
    ('IREU',  'Return from End User',                'end_user'),
    ('ISUP',  'In Supply Non-MSF',                   'third_party'),
    ('IBR',   'In Borrowing',                        'third_party'),
    ('IROL',  'In Return of Loan',                   'third_party'),
    ('ICPT',  'In Correction of Previous Transaction', None),
]

OUT_TYPES = [
    ('OEU',   'Issue to End User',    'end_user'),
    ('OEXP',  'Expired Items',        None),
    ('ODMG',  'Damaged Items',        None),
    ('OCCB',  'Cold Chain Break',     None),
    ('OBRC',  'Batch Recall',         None),
    ('OTHF',  'Theft',                None),
    ('OLS',   'Other Losses',         None),
    ('ODN',   'Out Donation',         'third_party'),
    ('OROB',  'Return of Borrowing',  'third_party'),
    ('OLOAN', 'Loan',                 'third_party'),
    ('OQRT',  'Quarantine',           None),
]

_IN_TYPE_MAP  = {code: (label, party) for code, label, party in IN_TYPES}
_OUT_TYPE_MAP = {code: (label, party) for code, label, party in OUT_TYPES}


def _mov_db():
    conn = get_db_connection()
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def generate_doc_number(conn, doc_type):
    """Generate next sequential document number: ABBR/YY/TYPE/001"""
    year = datetime.now().year
    yy   = str(year)[-2:]
    row  = conn.execute(
        "SELECT mission_abbreviation FROM mission_details WHERE is_active=1 LIMIT 1"
    ).fetchone()
    abbr = row['mission_abbreviation'] if row else 'MSF'
    conn.execute(
        "INSERT INTO doc_sequences(doc_type, year, last_seq) VALUES(?, ?, 1) "
        "ON CONFLICT(doc_type, year) DO UPDATE SET last_seq = last_seq + 1",
        (doc_type, year)
    )
    seq = conn.execute(
        "SELECT last_seq FROM doc_sequences WHERE doc_type=? AND year=?",
        (doc_type, year)
    ).fetchone()['last_seq']
    return f"{abbr}/{yy}/{doc_type}/{seq:03d}"


def _build_packing_list_excel(conn, movement_id):
    """Build openpyxl workbook for a movement packing list. Returns (wb, pl_number)."""
    mov = conn.execute('''
        SELECT m.*, u.username AS created_by_name,
               eu.name AS end_user_name,
               tp.name AS third_party_name,
               p_src.project_name AS source_project_name,
               p_dst.project_name AS dest_project_name
        FROM movements m
        LEFT JOIN users u      ON u.id = m.created_by
        LEFT JOIN end_users eu ON eu.end_user_id = m.end_user_id
        LEFT JOIN third_parties tp ON tp.third_party_id = m.third_party_id
        LEFT JOIN projects p_src ON p_src.project_code = m.source_project
        LEFT JOIN projects p_dst ON p_dst.project_code = m.dest_project
        WHERE m.id = ?
    ''', (movement_id,)).fetchone()

    lines = conn.execute(
        "SELECT * FROM movement_lines WHERE movement_id=? ORDER BY line_no, id",
        (movement_id,)
    ).fetchall()

    pl_number = generate_doc_number(conn, 'PL')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Packing List'

    # Page setup — landscape, fit all columns to one page width
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    title_blue  = "1F3A8A"
    header_fill = PatternFill(start_color=title_blue, end_color=title_blue, fill_type='solid')
    hdr_font    = Font(bold=True, color='FFFFFF', size=10)
    bold        = Font(bold=True)

    mission = conn.execute(
        "SELECT mission_name FROM mission_details WHERE is_active=1 LIMIT 1"
    ).fetchone()
    mission_name = mission['mission_name'] if mission else ''

    # ── Header block ──────────────────────────────────────────────────────
    ws['A1'] = 'PACKING LIST'
    ws['A1'].font = Font(bold=True, size=16, color=title_blue)
    ws.merge_cells('A1:M1')

    ws['A2'] = mission_name
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:G2')
    ws['H2'] = f"PL No: {pl_number}"
    ws['H2'].font = Font(bold=True, size=11)
    ws.merge_cells('H2:M2')

    ws['A3'] = f"Document No: {mov['document_number'] or ''}"
    ws.merge_cells('A3:D3')
    ws['E3'] = f"Date: {mov['movement_date'] or ''}"
    ws.merge_cells('E3:H3')
    ws['I3'] = f"Type: {mov['doc_type'] or ''}"
    ws.merge_cells('I3:M3')

    from_str = mov['source_project_name'] or mov['source_project'] or ''
    to_str   = (mov['end_user_name'] or mov['third_party_name'] or
                mov['dest_project_name'] or mov['dest_project'] or '')
    ws['A4'] = f"From: {from_str}"
    ws.merge_cells('A4:F4')
    ws['G4'] = f"To: {to_str}"
    ws.merge_cells('G4:M4')

    if mov['notes']:
        ws['A5'] = f"Notes: {mov['notes']}"
        ws.merge_cells('A5:M5')

    # ── Column headers ────────────────────────────────────────────────────
    HDR_ROW = 7
    headers = [
        ('#', 4), ('Item Code', 14), ('Description', 34), ('Batch No', 14),
        ('Exp Date', 11), ('Qty', 8), ('Unit', 8),
        ('Weight kg', 11), ('Volume m3', 11),
        ('Unit Price', 11), ('Currency', 10), ('Total Value', 13), ('Notes', 24),
    ]
    for col_idx, (hdr, width) in enumerate(headers, 1):
        cell = ws.cell(row=HDR_ROW, column=col_idx, value=hdr)
        cell.font      = hdr_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[HDR_ROW].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────
    total_weight = total_volume = total_value = 0.0
    alt_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')

    for i, ln in enumerate(lines, 1):
        r = HDR_ROW + i
        vals = [
            i, ln['item_code'] or '', ln['item_description'] or '',
            ln['batch_no'] or '', ln['exp_date'] or '',
            ln['qty'] or 0, ln['unit'] or '',
            ln['weight_kg'] or 0, ln['volume_m3'] or 0,
            ln['unit_price'] or 0, ln['currency'] or '', ln['total_value'] or 0,
            ln['notes'] or '',
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical='top', wrap_text=(c == 3))
            if i % 2 == 0:
                cell.fill = alt_fill
        total_weight += (ln['weight_kg']   or 0)
        total_volume += (ln['volume_m3']   or 0)
        total_value  += (ln['total_value'] or 0)

    # ── Totals row ────────────────────────────────────────────────────────
    tr = HDR_ROW + len(lines) + 1
    for c in range(1, 14):
        ws.cell(row=tr, column=c).fill = PatternFill(
            start_color='D0D8F0', end_color='D0D8F0', fill_type='solid')
    ws.cell(row=tr, column=1, value='TOTAL').font = bold
    ws.cell(row=tr, column=8, value=round(total_weight, 3)).font = bold
    ws.cell(row=tr, column=9, value=round(total_volume, 3)).font = bold
    ws.cell(row=tr, column=12, value=round(total_value, 2)).font = bold

    ws.freeze_panes = f'A{HDR_ROW + 1}'
    return wb, pl_number


# ── Certificate number generator ───────────────────────────────────────────
def generate_cert_number(conn, project_code, cert_type):
    """Generate YY/project_code/cert_type/seq  e.g. 26/CD502/DI/001"""
    year = datetime.now().year
    yy   = str(year)[-2:]
    key  = f"CERT_{cert_type}_{project_code}"
    conn.execute(
        "INSERT INTO doc_sequences(doc_type, year, last_seq) VALUES(?,?,1) "
        "ON CONFLICT(doc_type, year) DO UPDATE SET last_seq=last_seq+1",
        (key, year)
    )
    seq = conn.execute(
        "SELECT last_seq FROM doc_sequences WHERE doc_type=? AND year=?",
        (key, year)
    ).fetchone()['last_seq']
    return f"{yy}/{project_code}/{cert_type}/{seq:03d}"


# ── Donation / Loan certificate Excel builder ─────────────────────────────
def _build_cert_excel(conn, movement_id):
    """Build styled A4-portrait certificate for donation (IDN/ODN) or loan (IBR/IROL/OLOAN/OROB)."""
    CERT_TYPE_MAP = {
        'IDN': ('DI', 'DONATION CERTIFICATE'),
        'ODN': ('DO', 'DONATION CERTIFICATE'),
        'IBR': ('LI', 'LOAN RECEIPT'),
        'IROL':('LI', 'LOAN RECEIPT'),
        'OLOAN':('LO', 'LOAN CERTIFICATE'),
        'OROB': ('LO', 'LOAN CERTIFICATE'),
    }
    mov = conn.execute('''
        SELECT m.*, eu.name AS end_user_name, eu.address AS end_user_address,
               tp.name AS third_party_name, tp.address AS third_party_address,
               md.mission_name
        FROM movements m
        LEFT JOIN end_users eu    ON eu.end_user_id  = m.end_user_id
        LEFT JOIN third_parties tp ON tp.third_party_id = m.third_party_id
        LEFT JOIN mission_details md ON md.is_active = 1
        WHERE m.id = ?
    ''', (movement_id,)).fetchone()
    if not mov:
        raise ValueError(f"Movement {movement_id} not found")

    lines = conn.execute(
        "SELECT * FROM movement_lines WHERE movement_id=? ORDER BY line_no, id",
        (movement_id,)
    ).fetchall()

    doc_type = mov['doc_type']
    cert_key, cert_title = CERT_TYPE_MAP.get(doc_type, ('DI', 'CERTIFICATE'))
    project_code = mov['dest_project'] or mov['source_project'] or 'GEN'
    cert_num = generate_cert_number(conn, project_code, cert_key)

    party_name    = mov['third_party_name'] or mov['end_user_name'] or '—'
    party_addr    = mov['third_party_address'] or mov['end_user_address'] or ''
    mission_name  = mov['mission_name'] or ''

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Certificate'

    # A4 portrait, fit to one page
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize   = 9   # A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    title_blue   = "1F3A8A"
    hdr_fill     = PatternFill(start_color=title_blue, end_color=title_blue, fill_type='solid')
    hdr_font     = Font(bold=True, color='FFFFFF', size=10)
    bold         = Font(bold=True)
    bold14       = Font(bold=True, size=14, color=title_blue)
    thin         = Side(style='thin')
    thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')

    # Column widths
    col_widths = [4, 14, 34, 8, 8, 13, 11, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Row 1: mission left, cert number right ───────────────────────────
    ws.cell(row=1, column=1, value=mission_name).font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=7, value=f"No: {cert_num}").font = Font(bold=True, size=11)
    ws.merge_cells('G1:I1')
    ws.cell(row=1, column=7).alignment = Alignment(horizontal='right')

    # ── Row 2: certificate title ─────────────────────────────────────────
    ws.cell(row=2, column=1, value=cert_title).font = bold14
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:I2')

    # ── Row 3: date ──────────────────────────────────────────────────────
    ws.cell(row=3, column=1, value=f"Date: {mov['movement_date'] or ''}")
    ws.merge_cells('A3:I3')

    # ── Row 5: parties ────────────────────────────────────────────────────
    label_from = "Donor:" if doc_type in ('IDN','ODN') else "Lender:"
    label_to   = "Recipient:" if doc_type in ('IDN','ODN') else "Borrower:"
    ws.cell(row=5, column=1, value=f"{label_from} {party_name}").font = bold
    ws.merge_cells('A5:I5')
    if party_addr:
        ws.cell(row=6, column=1, value=f"Address: {party_addr}")
        ws.merge_cells('A6:I6')

    project_str = mov['dest_project'] or mov['source_project'] or ''
    ws.cell(row=7, column=1, value=f"{label_to} {mission_name} — Project: {project_str}").font = bold
    ws.merge_cells('A7:I7')

    # ── Row 9: table header ───────────────────────────────────────────────
    HDR_ROW = 9
    hdrs = ['#', 'Item Code', 'Description', 'Qty', 'Unit', 'Batch No', 'Exp Date', 'Unit Price', 'Total Value']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=HDR_ROW, column=c, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
    ws.row_dimensions[HDR_ROW].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────
    total_value = 0.0
    for i, ln in enumerate(lines, 1):
        r  = HDR_ROW + i
        tv = (ln['qty'] or 0) * (ln['unit_price'] or 0)
        total_value += tv
        vals = [i, ln['item_code'] or '', ln['item_description'] or '',
                ln['qty'] or 0, ln['unit'] or '',
                ln['batch_no'] or '', ln['exp_date'] or '',
                ln['unit_price'] or 0, tv]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(c == 3))
            if i % 2 == 0:
                cell.fill = alt_fill

    # ── Total row ─────────────────────────────────────────────────────────
    tr = HDR_ROW + len(lines) + 1
    total_fill = PatternFill(start_color='D0D8F0', end_color='D0D8F0', fill_type='solid')
    for c in range(1, 10):
        cell = ws.cell(row=tr, column=c)
        cell.fill   = total_fill
        cell.border = thin_border
    ws.cell(row=tr, column=1, value='TOTAL').font = bold
    ws.cell(row=tr, column=9, value=round(total_value, 2)).font = bold

    # ── Signature block ────────────────────────────────────────────────────
    sig_row = tr + 3
    ws.cell(row=sig_row,   column=1, value=f"{label_from.rstrip(':')} Signature:").font = bold
    ws.cell(row=sig_row,   column=6, value=f"{label_to.rstrip(':')} Signature:").font = bold
    ws.cell(row=sig_row+2, column=1, value='_' * 28)
    ws.cell(row=sig_row+2, column=6, value='_' * 28)
    ws.cell(row=sig_row+3, column=1, value='Name:')
    ws.cell(row=sig_row+3, column=6, value='Name:')
    ws.cell(row=sig_row+4, column=1, value='Date:')
    ws.cell(row=sig_row+4, column=6, value='Date:')

    ws.print_area = f'A1:I{sig_row+4}'
    return wb


# ── Packing list from raw lines (for dispatch without movement record) ─────
def _build_pl_from_lines(lines_data, header, conn):
    """Build PL Excel from a list of dicts (no movement record needed).
    header: {doc_type, source_project, dest_project, end_user_id, third_party_id,
             movement_date, notes}
    """
    eu = tp = None
    if header.get('end_user_id'):
        eu = conn.execute("SELECT name FROM end_users WHERE end_user_id=?",
                          (header['end_user_id'],)).fetchone()
    if header.get('third_party_id'):
        tp = conn.execute("SELECT name FROM third_parties WHERE third_party_id=?",
                          (header['third_party_id'],)).fetchone()
    mission = conn.execute(
        "SELECT mission_name FROM mission_details WHERE is_active=1 LIMIT 1"
    ).fetchone()
    mission_name = mission['mission_name'] if mission else ''

    pl_number = generate_doc_number(conn, 'PL')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Packing List'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    title_blue  = "1F3A8A"
    header_fill = PatternFill(start_color=title_blue, end_color=title_blue, fill_type='solid')
    hdr_font    = Font(bold=True, color='FFFFFF', size=10)
    bold        = Font(bold=True)
    alt_fill    = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')

    ws['A1'] = 'PACKING LIST'
    ws['A1'].font = Font(bold=True, size=16, color=title_blue)
    ws.merge_cells('A1:M1')
    ws['A2'] = mission_name
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:G2')
    ws['H2'] = f"PL No: {pl_number}"
    ws['H2'].font = Font(bold=True, size=11)
    ws.merge_cells('H2:M2')
    ws['A3'] = f"Type: {header.get('doc_type','')}"
    ws['E3'] = f"Date: {header.get('movement_date','')}"
    ws.merge_cells('A3:D3'); ws.merge_cells('E3:M3')
    to_str = (eu['name'] if eu else '') or (tp['name'] if tp else '') or header.get('dest_project','')
    ws['A4'] = f"From: {header.get('source_project','')}"
    ws['G4'] = f"To: {to_str}"
    ws.merge_cells('A4:F4'); ws.merge_cells('G4:M4')
    if header.get('notes'):
        ws['A5'] = f"Notes: {header['notes']}"
        ws.merge_cells('A5:M5')

    HDR_ROW = 7
    headers = [
        ('#', 4), ('Parcel No', 12), ('Item Code', 14), ('Description', 30),
        ('Batch No', 14), ('Exp Date', 11), ('Qty', 8), ('Unit', 8),
        ('Weight kg', 11), ('Volume m3', 11), ('Notes', 20),
    ]
    for col_idx, (hdr, width) in enumerate(headers, 1):
        cell = ws.cell(row=HDR_ROW, column=col_idx, value=hdr)
        cell.font = hdr_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[HDR_ROW].height = 22

    total_weight = total_volume = 0.0
    for i, ln in enumerate(lines_data, 1):
        r = HDR_ROW + i
        vals = [i, ln.get('parcel_number',''), ln.get('item_code',''),
                ln.get('item_description',''), ln.get('batch_no',''),
                ln.get('exp_date',''), ln.get('qty',0), ln.get('unit',''),
                ln.get('weight_kg',0), ln.get('volume_m3',0), ln.get('notes','')]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical='top', wrap_text=(c == 4))
            if i % 2 == 0:
                cell.fill = alt_fill
        total_weight += (ln.get('weight_kg') or 0)
        total_volume += (ln.get('volume_m3') or 0)

    tr = HDR_ROW + len(lines_data) + 1
    tot_fill = PatternFill(start_color='D0D8F0', end_color='D0D8F0', fill_type='solid')
    for c in range(1, 12):
        ws.cell(row=tr, column=c).fill = tot_fill
    ws.cell(row=tr, column=1, value='TOTAL').font = bold
    ws.cell(row=tr, column=9, value=round(total_weight, 3)).font = bold
    ws.cell(row=tr, column=10, value=round(total_volume, 3)).font = bold
    ws.freeze_panes = f'A{HDR_ROW + 1}'
    return wb, pl_number


# ── Movement types list ────────────────────────────────────────────────────
@app.route('/api/movements/types', methods=['GET'])
@login_required
def mov_types():
    direction = request.args.get('direction', '').upper()
    def fmt(types_list):
        return [{'code': c, 'label': l, 'required_party': p} for c, l, p in types_list]
    if direction == 'IN':
        return jsonify({'success': True, 'types': fmt(IN_TYPES)})
    if direction == 'OUT':
        return jsonify({'success': True, 'types': fmt(OUT_TYPES)})
    return jsonify({'success': True, 'in_types': fmt(IN_TYPES), 'out_types': fmt(OUT_TYPES)})


# ── IN movements list ─────────────────────────────────────────────────────
@app.route('/api/movements/in', methods=['GET'])
@login_required
def mov_in_list():
    conn = None
    try:
        conn = _mov_db()
        rows = conn.execute('''
            SELECT m.id, m.document_number, m.doc_type, m.movement_date,
                   m.dest_project, m.status, m.total_weight_kg, m.total_volume_m3,
                   m.notes, m.created_at,
                   u.username  AS created_by_name,
                   eu.name     AS end_user_name,
                   tp.name     AS third_party_name,
                   (SELECT COUNT(*) FROM movement_lines ml WHERE ml.movement_id = m.id) AS line_count
            FROM movements m
            LEFT JOIN users u      ON u.id = m.created_by
            LEFT JOIN end_users eu ON eu.end_user_id = m.end_user_id
            LEFT JOIN third_parties tp ON tp.third_party_id = m.third_party_id
            WHERE m.movement_type = 'IN'
            ORDER BY m.created_at DESC
        ''').fetchall()
        return jsonify({'success': True, 'movements': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── IN movement detail ────────────────────────────────────────────────────
@app.route('/api/movements/in/<int:mov_id>', methods=['GET'])
@login_required
def mov_in_get(mov_id):
    conn = None
    try:
        conn = _mov_db()
        mov = conn.execute('''
            SELECT m.*, u.username AS created_by_name,
                   eu.name AS end_user_name, tp.name AS third_party_name
            FROM movements m
            LEFT JOIN users u      ON u.id = m.created_by
            LEFT JOIN end_users eu ON eu.end_user_id = m.end_user_id
            LEFT JOIN third_parties tp ON tp.third_party_id = m.third_party_id
            WHERE m.id = ? AND m.movement_type = 'IN'
        ''', (mov_id,)).fetchone()
        if not mov:
            return jsonify({'success': False, 'message': 'Not found'}), 404
        lines = conn.execute(
            "SELECT * FROM movement_lines WHERE movement_id=? ORDER BY line_no, id",
            (mov_id,)
        ).fetchall()
        return jsonify({'success': True, 'movement': dict(mov),
                        'lines': [dict(l) for l in lines]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Create / Update IN movement (Draft) ──────────────────────────────────
@app.route('/api/movements/in', methods=['POST'])
@login_required
def mov_in_save():
    conn = None
    try:
        data  = request.get_json()
        mov_id = data.get('id')
        conn  = _mov_db()

        doc_type = data.get('doc_type', '')
        if doc_type not in _IN_TYPE_MAP:
            return jsonify({'success': False, 'message': 'Invalid IN type'}), 400

        # Validate required party fields
        _, req_party = _IN_TYPE_MAP[doc_type]
        end_user_id    = data.get('end_user_id') or None
        third_party_id = data.get('third_party_id') or None
        if req_party == 'end_user'    and not end_user_id:
            return jsonify({'success': False,
                            'message': f'End User is required for {doc_type}'}), 400
        if req_party == 'third_party' and not third_party_id:
            return jsonify({'success': False,
                            'message': f'Third Party is required for {doc_type}'}), 400

        lines  = data.get('lines', [])
        tw, tv_m3 = 0.0, 0.0
        for ln in lines:
            tw   += float(ln.get('weight_kg') or 0)
            tv_m3 += float(ln.get('volume_m3') or 0)

        if mov_id:
            # Update existing Draft
            conn.execute('''
                UPDATE movements SET
                    doc_type=?, movement_date=?, dest_project=?,
                    end_user_id=?, third_party_id=?,
                    total_weight_kg=?, total_volume_m3=?, notes=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=? AND movement_type='IN' AND status='Draft'
            ''', (doc_type, data.get('movement_date'), data.get('dest_project'),
                  end_user_id, third_party_id, tw, tv_m3,
                  data.get('notes'), mov_id))
            conn.execute("DELETE FROM movement_lines WHERE movement_id=?", (mov_id,))
        else:
            cur = conn.execute('''
                INSERT INTO movements
                    (movement_type, doc_type, movement_date, dest_project,
                     end_user_id, third_party_id, total_weight_kg, total_volume_m3,
                     notes, created_by, status)
                VALUES ('IN', ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Draft')
            ''', (doc_type, data.get('movement_date'), data.get('dest_project'),
                  end_user_id, third_party_id, tw, tv_m3,
                  data.get('notes'), current_user.id))
            mov_id = cur.lastrowid

        # Insert lines
        for idx, ln in enumerate(lines, 1):
            qty        = float(ln.get('qty') or 0)
            unit_price = float(ln.get('unit_price') or 0)
            conn.execute('''
                INSERT INTO movement_lines
                    (movement_id, line_no, item_code, item_description,
                     qty, unit, batch_no, exp_date,
                     unit_price, currency, total_value,
                     weight_kg, volume_m3, pallet_number, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (mov_id, idx, ln.get('item_code'), ln.get('item_description'),
                  qty, ln.get('unit'), ln.get('batch_no'), ln.get('exp_date'),
                  unit_price, ln.get('currency', 'USD'), qty * unit_price,
                  float(ln.get('weight_kg') or 0), float(ln.get('volume_m3') or 0),
                  ln.get('pallet_number'), ln.get('notes')))

        conn.commit()
        return jsonify({'success': True, 'id': mov_id})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Delete IN Draft ───────────────────────────────────────────────────────
@app.route('/api/movements/in/<int:mov_id>', methods=['DELETE'])
@login_required
def mov_in_delete(mov_id):
    conn = None
    try:
        conn = _mov_db()
        conn.execute(
            "DELETE FROM movements WHERE id=? AND movement_type='IN' AND status='Draft'",
            (mov_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Confirm IN movement ───────────────────────────────────────────────────
@app.route('/api/movements/in/<int:mov_id>/confirm', methods=['POST'])
@login_required
def mov_in_confirm(mov_id):
    conn = None
    try:
        conn = _mov_db()
        mov = conn.execute(
            "SELECT * FROM movements WHERE id=? AND movement_type='IN' AND status='Draft'",
            (mov_id,)).fetchone()
        if not mov:
            return jsonify({'success': False, 'message': 'Draft not found'}), 404

        lines = conn.execute(
            "SELECT * FROM movement_lines WHERE movement_id=?", (mov_id,)
        ).fetchall()
        if not lines:
            return jsonify({'success': False,
                            'message': 'Cannot confirm: no lines added'}), 400

        doc_num = generate_doc_number(conn, mov['doc_type'])
        conn.execute(
            "UPDATE movements SET document_number=?, status='Confirmed', "
            "updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (doc_num, mov_id))
        conn.execute(
            "UPDATE movement_lines SET document_number=? WHERE movement_id=?",
            (doc_num, mov_id))
        conn.commit()
        return jsonify({'success': True, 'document_number': doc_num})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Export IN packing list ────────────────────────────────────────────────
@app.route('/api/movements/in/<int:mov_id>/export', methods=['GET'])
@login_required
def mov_in_export(mov_id):
    conn = None
    try:
        conn = _mov_db()
        if not conn.execute(
            "SELECT id FROM movements WHERE id=? AND movement_type='IN'",
            (mov_id,)).fetchone():
            return jsonify({'success': False, 'message': 'Not found'}), 404

        wb, pl_num = _build_packing_list_excel(conn, mov_id)
        conn.commit()  # save PL sequence increment
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_pl = pl_num.replace('/', '_')
        return send_file(buf, as_attachment=True,
                         download_name=f"PL_{safe_pl}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════
#  OUT MOVEMENTS
# ═══════════════════════════════════════════════════════════════════════════

def _available_stock_query(conn, project_code):
    """
    Returns rows: item_code, item_description, batch_no, exp_date,
                  project_code, available_qty
    Ordered by exp_date ASC (FEFO).
    Combines: cargo receptions + confirmed IN movements - confirmed OUT movements.
    """
    rows = conn.execute('''
        WITH receptions AS (
            SELECT item_code,
                   MAX(item_description) AS item_description,
                   COALESCE(batch_no,'') AS batch_no,
                   COALESCE(exp_date,'') AS exp_date,
                   project_code,
                   SUM(qty_received) AS qty
            FROM stock_transactions
            WHERE transaction_type = 'RECEPTION'
              AND project_code = ?
            GROUP BY item_code, batch_no, exp_date, project_code
        ),
        in_mvts AS (
            SELECT ml.item_code,
                   MAX(ml.item_description) AS item_description,
                   COALESCE(ml.batch_no,'') AS batch_no,
                   COALESCE(ml.exp_date,'') AS exp_date,
                   m.dest_project AS project_code,
                   SUM(ml.qty) AS qty
            FROM movement_lines ml
            JOIN movements m ON m.id = ml.movement_id
            WHERE m.movement_type = 'IN' AND m.status = 'Confirmed'
              AND m.dest_project = ?
            GROUP BY ml.item_code, ml.batch_no, ml.exp_date, m.dest_project
        ),
        out_mvts AS (
            SELECT ml.item_code,
                   COALESCE(ml.batch_no,'') AS batch_no,
                   COALESCE(ml.exp_date,'') AS exp_date,
                   m.source_project AS project_code,
                   SUM(ml.qty) AS qty
            FROM movement_lines ml
            JOIN movements m ON m.id = ml.movement_id
            WHERE m.movement_type = 'OUT' AND m.status = 'Confirmed'
              AND m.source_project = ?
            GROUP BY ml.item_code, ml.batch_no, ml.exp_date, m.source_project
        ),
        all_in AS (
            SELECT * FROM receptions
            UNION ALL
            SELECT * FROM in_mvts
        ),
        total_in AS (
            SELECT item_code, MAX(item_description) AS item_description,
                   batch_no, exp_date, project_code, SUM(qty) AS qty
            FROM all_in
            GROUP BY item_code, batch_no, exp_date, project_code
        )
        SELECT ti.item_code, ti.item_description, ti.batch_no, ti.exp_date,
               ti.project_code,
               ti.qty - COALESCE(om.qty, 0) AS available_qty
        FROM total_in ti
        LEFT JOIN out_mvts om
               ON om.item_code    = ti.item_code
              AND om.batch_no     = ti.batch_no
              AND om.exp_date     = ti.exp_date
              AND om.project_code = ti.project_code
        WHERE ti.qty - COALESCE(om.qty, 0) > 0
        ORDER BY
            CASE WHEN ti.exp_date = '' OR ti.exp_date IS NULL THEN 1 ELSE 0 END,
            ti.exp_date ASC,
            ti.item_code ASC
    ''', (project_code, project_code, project_code)).fetchall()
    return rows


@app.route('/api/movements/stock', methods=['GET'])
@login_required
def mov_stock():
    """Available stock by project (FEFO sorted). ?project=CODE"""
    project = request.args.get('project', '').strip()
    if not project:
        return jsonify({'success': False, 'message': 'project param required'}), 400
    conn = None
    try:
        conn = _mov_db()
        rows = _available_stock_query(conn, project)
        return jsonify({'success': True, 'stock': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/movements/out', methods=['GET'])
@login_required
def mov_out_list():
    conn = None
    try:
        conn = _mov_db()
        rows = conn.execute('''
            SELECT m.id, m.document_number, m.doc_type, m.movement_date,
                   m.source_project, m.dest_project, m.status,
                   m.total_weight_kg, m.total_volume_m3, m.notes, m.created_at,
                   u.username  AS created_by_name,
                   eu.name     AS end_user_name,
                   tp.name     AS third_party_name,
                   (SELECT COUNT(*) FROM movement_lines ml WHERE ml.movement_id=m.id) AS line_count
            FROM movements m
            LEFT JOIN users u      ON u.id = m.created_by
            LEFT JOIN end_users eu ON eu.end_user_id = m.end_user_id
            LEFT JOIN third_parties tp ON tp.third_party_id = m.third_party_id
            WHERE m.movement_type = 'OUT'
            ORDER BY m.created_at DESC
        ''').fetchall()
        return jsonify({'success': True, 'movements': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/movements/out/<int:mov_id>', methods=['GET'])
@login_required
def mov_out_get(mov_id):
    conn = None
    try:
        conn = _mov_db()
        mov = conn.execute('''
            SELECT m.*, u.username AS created_by_name,
                   eu.name AS end_user_name, tp.name AS third_party_name
            FROM movements m
            LEFT JOIN users u      ON u.id = m.created_by
            LEFT JOIN end_users eu ON eu.end_user_id = m.end_user_id
            LEFT JOIN third_parties tp ON tp.third_party_id = m.third_party_id
            WHERE m.id=? AND m.movement_type='OUT'
        ''', (mov_id,)).fetchone()
        if not mov:
            return jsonify({'success': False, 'message': 'Not found'}), 404
        lines = conn.execute(
            "SELECT * FROM movement_lines WHERE movement_id=? ORDER BY line_no, id",
            (mov_id,)).fetchall()
        return jsonify({'success': True, 'movement': dict(mov),
                        'lines': [dict(l) for l in lines]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/movements/out', methods=['POST'])
@login_required
def mov_out_save():
    conn = None
    try:
        data   = request.get_json()
        mov_id = data.get('id')
        conn   = _mov_db()

        doc_type = data.get('doc_type', '')
        if doc_type not in _OUT_TYPE_MAP:
            return jsonify({'success': False, 'message': 'Invalid OUT type'}), 400

        _, req_party   = _OUT_TYPE_MAP[doc_type]
        end_user_id    = data.get('end_user_id') or None
        third_party_id = data.get('third_party_id') or None
        source_project = data.get('source_project', '').strip()

        if not source_project:
            return jsonify({'success': False, 'message': 'Source project is required'}), 400
        if req_party == 'end_user'    and not end_user_id:
            return jsonify({'success': False,
                            'message': f'End User is required for {doc_type}'}), 400
        if req_party == 'third_party' and not third_party_id:
            return jsonify({'success': False,
                            'message': f'Third Party is required for {doc_type}'}), 400

        lines = data.get('lines', [])
        tw, tv_m3 = 0.0, 0.0
        for ln in lines:
            tw    += float(ln.get('weight_kg') or 0)
            tv_m3 += float(ln.get('volume_m3') or 0)

        if mov_id:
            conn.execute('''
                UPDATE movements SET
                    doc_type=?, movement_date=?, source_project=?, dest_project=?,
                    end_user_id=?, third_party_id=?,
                    total_weight_kg=?, total_volume_m3=?, notes=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=? AND movement_type='OUT' AND status='Draft'
            ''', (doc_type, data.get('movement_date'), source_project,
                  data.get('dest_project') or None,
                  end_user_id, third_party_id, tw, tv_m3,
                  data.get('notes'), mov_id))
            conn.execute("DELETE FROM movement_lines WHERE movement_id=?", (mov_id,))
        else:
            cur = conn.execute('''
                INSERT INTO movements
                    (movement_type, doc_type, movement_date,
                     source_project, dest_project,
                     end_user_id, third_party_id,
                     total_weight_kg, total_volume_m3,
                     notes, created_by, status)
                VALUES ('OUT', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Draft')
            ''', (doc_type, data.get('movement_date'), source_project,
                  data.get('dest_project') or None,
                  end_user_id, third_party_id, tw, tv_m3,
                  data.get('notes'), current_user.id))
            mov_id = cur.lastrowid

        for idx, ln in enumerate(lines, 1):
            qty        = float(ln.get('qty') or 0)
            unit_price = float(ln.get('unit_price') or 0)
            conn.execute('''
                INSERT INTO movement_lines
                    (movement_id, line_no, item_code, item_description,
                     qty, unit, batch_no, exp_date,
                     unit_price, currency, total_value,
                     weight_kg, volume_m3, parcel_number, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (mov_id, idx, ln.get('item_code'), ln.get('item_description'),
                  qty, ln.get('unit'), ln.get('batch_no'), ln.get('exp_date'),
                  unit_price, ln.get('currency', 'USD'), qty * unit_price,
                  float(ln.get('weight_kg') or 0), float(ln.get('volume_m3') or 0),
                  ln.get('parcel_number'), ln.get('notes')))

        conn.commit()
        return jsonify({'success': True, 'id': mov_id})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/movements/out/<int:mov_id>', methods=['DELETE'])
@login_required
def mov_out_delete(mov_id):
    conn = None
    try:
        conn = _mov_db()
        conn.execute(
            "DELETE FROM movements WHERE id=? AND movement_type='OUT' AND status='Draft'",
            (mov_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/movements/out/<int:mov_id>/confirm', methods=['POST'])
@login_required
def mov_out_confirm(mov_id):
    conn = None
    try:
        conn = _mov_db()
        mov = conn.execute(
            "SELECT * FROM movements WHERE id=? AND movement_type='OUT' AND status='Draft'",
            (mov_id,)).fetchone()
        if not mov:
            return jsonify({'success': False, 'message': 'Draft not found'}), 404

        lines = conn.execute(
            "SELECT * FROM movement_lines WHERE movement_id=?", (mov_id,)
        ).fetchall()
        if not lines:
            return jsonify({'success': False,
                            'message': 'Cannot confirm: no lines added'}), 400

        # Server-side stock validation
        project = mov['source_project']
        avail   = {
            (r['item_code'], r['batch_no'] or '', r['exp_date'] or ''): r['available_qty']
            for r in _available_stock_query(conn, project)
        }
        for ln in lines:
            key   = (ln['item_code'], ln['batch_no'] or '', ln['exp_date'] or '')
            stock = avail.get(key, 0)
            if (ln['qty'] or 0) > stock:
                return jsonify({
                    'success': False,
                    'message': (f"Insufficient stock for {ln['item_code']} "
                                f"(batch: {ln['batch_no'] or '-'}, "
                                f"exp: {ln['exp_date'] or '-'}). "
                                f"Available: {stock}, Requested: {ln['qty']}")
                }), 400

        doc_num = generate_doc_number(conn, mov['doc_type'])
        conn.execute(
            "UPDATE movements SET document_number=?, status='Confirmed', "
            "updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (doc_num, mov_id))
        conn.execute(
            "UPDATE movement_lines SET document_number=? WHERE movement_id=?",
            (doc_num, mov_id))
        conn.commit()
        return jsonify({'success': True, 'document_number': doc_num})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/movements/out/<int:mov_id>/export', methods=['GET'])
@login_required
def mov_out_export(mov_id):
    conn = None
    try:
        conn = _mov_db()
        if not conn.execute(
            "SELECT id FROM movements WHERE id=? AND movement_type='OUT'",
            (mov_id,)).fetchone():
            return jsonify({'success': False, 'message': 'Not found'}), 404

        wb, pl_num = _build_packing_list_excel(conn, mov_id)
        conn.commit()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_pl = pl_num.replace('/', '_')
        return send_file(buf, as_attachment=True,
                         download_name=f"PL_{safe_pl}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Certificate routes ─────────────────────────────────────────────────────
@app.route('/api/movements/in/<int:mov_id>/certificate', methods=['GET'])
@login_required
def mov_in_certificate(mov_id):
    conn = None
    try:
        conn = _mov_db()
        if not conn.execute(
            "SELECT id FROM movements WHERE id=? AND movement_type='IN'",
            (mov_id,)).fetchone():
            return jsonify({'success': False, 'message': 'Not found'}), 404
        wb = _build_cert_excel(conn, mov_id)
        conn.commit()
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f"Certificate_{mov_id}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/movements/out/<int:mov_id>/certificate', methods=['GET'])
@login_required
def mov_out_certificate(mov_id):
    conn = None
    try:
        conn = _mov_db()
        if not conn.execute(
            "SELECT id FROM movements WHERE id=? AND movement_type='OUT'",
            (mov_id,)).fetchone():
            return jsonify({'success': False, 'message': 'Not found'}), 404
        wb = _build_cert_excel(conn, mov_id)
        conn.commit()
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f"Certificate_{mov_id}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Dispatch packing list from raw lines (no movement saved) ───────────────
@app.route('/api/dispatch/packing-list', methods=['POST'])
@login_required
def dispatch_packing_list():
    conn = None
    try:
        req    = request.get_json()
        lines  = req.get('lines', [])
        header = {k: req.get(k) for k in ('doc_type','source_project','dest_project',
                                            'end_user_id','third_party_id',
                                            'movement_date','notes')}
        if not lines:
            return jsonify({'success': False, 'message': 'No lines provided'}), 400
        conn = _mov_db()
        wb, pl_num = _build_pl_from_lines(lines, header, conn)
        conn.commit()
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        safe = pl_num.replace('/', '_')
        return send_file(buf, as_attachment=True,
                         download_name=f"DispatchPL_{safe}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Dispatch: received items (FEFO sorted for parcel-based OUT) ────────────
@app.route('/api/dispatch/items', methods=['GET'])
@login_required
def dispatch_items():
    project    = request.args.get('project') or None
    item_q     = request.args.get('item') or None
    parcel_q   = request.args.get('parcel') or None
    cargo_q    = request.args.get('cargo') or None
    conn = None
    try:
        conn = _cr_db()
        q  = '''
            SELECT parcel_number, item_code, item_description,
                   COALESCE(batch_no_received, batch_no)       AS batch_no,
                   COALESCE(exp_date_received, exp_date)       AS exp_date,
                   qty_unit_tot AS qty, packaging AS unit,
                   weight_kg, volume_m3, pallet_number, project_code,
                   packing_ref, field_ref, cargo_session_id, received_at
            FROM basic_data
            WHERE reception_number IS NOT NULL AND parcel_number IS NOT NULL
        '''
        params = []
        if project:
            q += ' AND project_code=?';              params.append(project)
        if item_q:
            q += ' AND (item_code LIKE ? OR item_description LIKE ?)';
            params += [f'%{item_q}%', f'%{item_q}%']
        if parcel_q:
            q += ' AND parcel_number LIKE ?';        params.append(f'%{parcel_q}%')
        if cargo_q:
            q += ' AND cargo_session_id LIKE ?';     params.append(f'%{cargo_q}%')
        q += ''' ORDER BY project_code ASC,
                          COALESCE(exp_date_received, exp_date) ASC,
                          parcel_number ASC'''
        rows = conn.execute(q, params).fetchall()
        return jsonify({'success': True, 'items': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Dispatch: single parcel contents ──────────────────────────────────────
@app.route('/api/dispatch/parcel/<parcel_number>', methods=['GET'])
@login_required
def dispatch_parcel_contents(parcel_number):
    conn = None
    try:
        conn = _cr_db()
        rows = conn.execute(
            '''SELECT parcel_number, item_code, item_description,
                      COALESCE(batch_no_received, batch_no) AS batch_no,
                      COALESCE(exp_date_received, exp_date) AS exp_date,
                      qty_unit_tot AS qty, packaging AS unit,
                      weight_kg, volume_m3, pallet_number, project_code,
                      packing_ref, field_ref, cargo_session_id
               FROM basic_data
               WHERE parcel_number=? AND reception_number IS NOT NULL
               ORDER BY line_no''',
            (parcel_number,)
        ).fetchall()
        return jsonify({'success': True, 'items': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Dispatch parcel map — all parcels with status ────────────────────────
@app.route('/api/dispatch/parcel-map', methods=['GET'])
@login_required
def dispatch_parcel_map():
    project = request.args.get('project') or None
    search  = request.args.get('search')  or None
    conn = None
    try:
        conn = get_db_connection()
        # Get set of dispatched parcel numbers (in confirmed OUT movements)
        try:
            disp_rows = conn.execute(
                '''SELECT DISTINCT ml.parcel_number FROM movement_lines ml
                   JOIN movements m ON ml.movement_id = m.id
                   WHERE m.movement_type='OUT' AND m.status='Confirmed'
                     AND ml.parcel_number IS NOT NULL AND ml.parcel_number != ""'''
            ).fetchall()
            dispatched_parcels = {r['parcel_number'] for r in disp_rows}
        except Exception:
            dispatched_parcels = set()

        q = '''
            SELECT parcel_number, project_code, packing_ref, pallet_number, order_type,
                   field_ref, COUNT(*) AS item_count,
                   SUM(weight_kg) AS total_weight,
                   MAX(reception_number) AS reception_number
            FROM basic_data
            WHERE parcel_number IS NOT NULL AND parcel_number != ""
        '''
        params = []
        if project:
            q += ' AND project_code = ?'; params.append(project)
        if search:
            q += ' AND (parcel_number LIKE ? OR packing_ref LIKE ?)'; params += [f'%{search}%', f'%{search}%']
        q += ' GROUP BY parcel_number ORDER BY project_code ASC, packing_ref ASC, CAST(parcel_number AS INTEGER) ASC'

        rows = conn.execute(q, params).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            if r['parcel_number'] in dispatched_parcels:
                d['status'] = 'dispatched'
            elif r['reception_number']:
                d['status'] = 'received'
            else:
                d['status'] = 'pending'
            result.append(d)
        return jsonify({'success': True, 'parcels': result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Reception Report data ─────────────────────────────────────────────────
@app.route('/api/reports/reception', methods=['GET'])
@login_required
def reception_report():
    project    = request.args.get('project') or None
    order_type = request.args.get('order_type') or None
    rec_no     = request.args.get('reception_number') or None
    cargo_id   = request.args.get('cargo_session') or None
    date_from  = request.args.get('date_from') or None
    date_to    = request.args.get('date_to') or None
    conn = None
    try:
        conn = _cr_db()
        q = '''
            SELECT b.parcel_number, b.field_ref, b.project_code, b.order_type,
                   MAX(b.pallet_number) AS pallet_number,
                   COUNT(*) AS item_count,
                   SUM(b.weight_kg) AS total_weight,
                   SUM(b.volume_m3) AS total_volume,
                   MAX(b.reception_number) AS reception_number,
                   MAX(b.received_at)      AS received_at,
                   MAX(b.cargo_session_id) AS cargo_session_id,
                   MAX(b.parcel_note)      AS parcel_note,
                   u.username AS received_by_name
            FROM basic_data b
            LEFT JOIN users u ON u.id = b.received_by
            WHERE b.reception_number IS NOT NULL AND b.parcel_number IS NOT NULL
        '''
        params = []
        if project:
            q += ' AND b.project_code=?';                 params.append(project)
        if order_type:
            q += ' AND b.order_type=?';                   params.append(order_type)
        if rec_no:
            q += ' AND b.reception_number LIKE ?';        params.append(f'%{rec_no}%')
        if cargo_id:
            q += ' AND b.cargo_session_id LIKE ?';        params.append(f'%{cargo_id}%')
        if date_from:
            q += ' AND DATE(b.received_at) >= ?';         params.append(date_from)
        if date_to:
            q += ' AND DATE(b.received_at) <= ?';         params.append(date_to)
        q += ' GROUP BY b.parcel_number ORDER BY b.received_at DESC'
        rows = conn.execute(q, params).fetchall()
        # Summary
        total_parcels = len(rows)
        total_items   = sum(r['item_count'] or 0 for r in rows)
        total_weight  = sum(r['total_weight'] or 0 for r in rows)
        return jsonify({'success': True,
                        'rows': [dict(r) for r in rows],
                        'summary': {'total_parcels': total_parcels,
                                    'total_items': total_items,
                                    'total_weight': round(total_weight, 2)}})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Reception Report Excel export ─────────────────────────────────────────
@app.route('/api/reports/reception/export', methods=['GET'])
@login_required
def reception_report_export():
    conn = None
    try:
        conn = _cr_db()
        # Re-use same query logic as /api/reports/reception
        project    = request.args.get('project') or None
        order_type = request.args.get('order_type') or None
        rec_no     = request.args.get('reception_number') or None
        cargo_id   = request.args.get('cargo_session') or None
        date_from  = request.args.get('date_from') or None
        date_to    = request.args.get('date_to') or None

        q = '''
            SELECT b.parcel_number, b.field_ref, b.project_code, b.order_type,
                   MAX(b.pallet_number) AS pallet_number,
                   COUNT(*) AS item_count,
                   SUM(b.weight_kg) AS total_weight,
                   SUM(b.volume_m3) AS total_volume,
                   MAX(b.reception_number) AS reception_number,
                   MAX(b.received_at)      AS received_at,
                   MAX(b.cargo_session_id) AS cargo_session_id,
                   MAX(b.parcel_note)      AS parcel_note,
                   u.username AS received_by_name
            FROM basic_data b
            LEFT JOIN users u ON u.id = b.received_by
            WHERE b.reception_number IS NOT NULL AND b.parcel_number IS NOT NULL
        '''
        params = []
        if project:    q += ' AND b.project_code=?';         params.append(project)
        if order_type: q += ' AND b.order_type=?';           params.append(order_type)
        if rec_no:     q += ' AND b.reception_number LIKE ?'; params.append(f'%{rec_no}%')
        if cargo_id:   q += ' AND b.cargo_session_id LIKE ?'; params.append(f'%{cargo_id}%')
        if date_from:  q += ' AND DATE(b.received_at) >= ?';  params.append(date_from)
        if date_to:    q += ' AND DATE(b.received_at) <= ?';  params.append(date_to)
        q += ' GROUP BY b.parcel_number ORDER BY b.received_at DESC'
        rows = conn.execute(q, params).fetchall()

        mission = conn.execute(
            "SELECT mission_name FROM mission_details WHERE is_active=1 LIMIT 1"
        ).fetchone()
        mission_name = mission['mission_name'] if mission else ''

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Reception Report'
        title_blue  = "1F3A8A"
        hdr_fill    = PatternFill(start_color=title_blue, end_color=title_blue, fill_type='solid')
        hdr_font    = Font(bold=True, color='FFFFFF', size=10)
        alt_fill    = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')

        ws['A1'] = f"{mission_name} — Reception Report"
        ws['A1'].font = Font(bold=True, size=13, color=title_blue)
        ws.merge_cells('A1:M1')
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells('A2:M2')

        HDR = 4
        hdrs = [('#','#',4),('Parcel No','parcel_number',14),('Field Ref','field_ref',22),
                ('Project','project_code',12),('Type','order_type',13),
                ('Pallet','pallet_number',10),('Items','item_count',7),
                ('Weight kg','total_weight',11),('Volume m3','total_volume',11),
                ('Reception No','reception_number',18),('Received At','received_at',18),
                ('Received By','received_by_name',16),('Notes','parcel_note',22)]
        for c,(lbl,_,w) in enumerate(hdrs,1):
            cell = ws.cell(row=HDR, column=c, value=lbl)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
        ws.row_dimensions[HDR].height = 18

        for i, row in enumerate(rows, 1):
            r = HDR + i
            vals = [i, row['parcel_number'], row['field_ref'], row['project_code'],
                    row['order_type'], row['pallet_number'], row['item_count'],
                    round(row['total_weight'] or 0, 2), round(row['total_volume'] or 0, 3),
                    row['reception_number'], str(row['received_at'] or '')[:16],
                    row['received_by_name'], row['parcel_note']]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = Alignment(vertical='top')
                if i % 2 == 0: cell.fill = alt_fill

        ws.freeze_panes = f'A{HDR+1}'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"ReceptionReport_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════
#  REPORTS  (Phase 4)
# ═══════════════════════════════════════════════════════════════════════════

def _reports_db():
    return _mov_db()


def _stock_summary_rows(conn, project=None, item_filter=None):
    """
    Returns net stock per (project_code, item_code, batch_no, exp_date).
    Sources: cargo receptions + IN movements - OUT movements.
    bd_receptions fallback covers parcels received before stock_transactions was used.
    """
    params_r = []
    params_b = []   # fallback: basic_data for parcels not in stock_transactions
    params_i = []
    params_o = []

    where_r = "WHERE st.transaction_type = 'RECEPTION'"
    where_b = ("WHERE bd.reception_number IS NOT NULL "
               "AND bd.item_code IS NOT NULL "
               "AND bd.qty_unit_tot IS NOT NULL AND bd.qty_unit_tot > 0")
    where_i = "WHERE m.movement_type='IN' AND m.status='Confirmed'"
    where_o = "WHERE m.movement_type='OUT' AND m.status='Confirmed'"

    if project:
        where_r += " AND st.project_code = ?"
        where_b += " AND bd.project_code = ?"
        where_i += " AND m.dest_project = ?"
        where_o += " AND m.source_project = ?"
        params_r.append(project)
        params_b.append(project)
        params_i.append(project)
        params_o.append(project)

    if item_filter:
        like = f"%{item_filter}%"
        where_r += " AND (st.item_code LIKE ? OR st.item_description LIKE ?)"
        where_b += " AND (bd.item_code LIKE ? OR bd.item_description LIKE ?)"
        where_i += " AND (ml.item_code LIKE ? OR ml.item_description LIKE ?)"
        where_o += " AND ml.item_code LIKE ?"
        params_r += [like, like]
        params_b += [like, like]
        params_i += [like, like]
        params_o.append(like)

    sql = f'''
        WITH receptions AS (
            SELECT st.project_code AS project_code,
                   st.item_code, MAX(st.item_description) AS item_description,
                   COALESCE(st.batch_no,'') AS batch_no,
                   COALESCE(st.exp_date,'') AS exp_date,
                   SUM(st.qty_received) AS qty_in, 0.0 AS qty_out
            FROM stock_transactions st
            {where_r}
            GROUP BY st.project_code, st.item_code, st.batch_no, st.exp_date
        ),
        bd_receptions AS (
            -- Fallback: count received basic_data rows whose parcel is not yet in stock_transactions
            SELECT bd.project_code AS project_code,
                   bd.item_code, MAX(bd.item_description) AS item_description,
                   COALESCE(bd.batch_no_received, bd.batch_no,'') AS batch_no,
                   COALESCE(bd.exp_date_received, bd.exp_date,'') AS exp_date,
                   SUM(bd.qty_unit_tot) AS qty_in, 0.0 AS qty_out
            FROM basic_data bd
            {where_b}
              AND (bd.parcel_number IS NULL
                   OR bd.parcel_number NOT IN (
                       SELECT DISTINCT parcel_number FROM stock_transactions
                       WHERE transaction_type='RECEPTION' AND parcel_number IS NOT NULL))
            GROUP BY bd.project_code, bd.item_code, bd.batch_no, bd.exp_date
        ),
        in_mvts AS (
            SELECT m.dest_project AS project_code,
                   ml.item_code, MAX(ml.item_description) AS item_description,
                   COALESCE(ml.batch_no,'') AS batch_no,
                   COALESCE(ml.exp_date,'') AS exp_date,
                   SUM(ml.qty) AS qty_in, 0.0 AS qty_out
            FROM movement_lines ml JOIN movements m ON m.id=ml.movement_id
            {where_i}
            GROUP BY m.dest_project, ml.item_code, ml.batch_no, ml.exp_date
        ),
        out_mvts AS (
            SELECT m.source_project AS project_code,
                   ml.item_code, '' AS item_description,
                   COALESCE(ml.batch_no,'') AS batch_no,
                   COALESCE(ml.exp_date,'') AS exp_date,
                   0.0 AS qty_in, SUM(ml.qty) AS qty_out
            FROM movement_lines ml JOIN movements m ON m.id=ml.movement_id
            {where_o}
            GROUP BY m.source_project, ml.item_code, ml.batch_no, ml.exp_date
        ),
        combined AS (
            SELECT * FROM receptions
            UNION ALL SELECT * FROM bd_receptions
            UNION ALL SELECT * FROM in_mvts
            UNION ALL SELECT * FROM out_mvts
        )
        SELECT project_code, item_code, MAX(item_description) AS item_description,
               batch_no, exp_date,
               SUM(qty_in) AS total_in, SUM(qty_out) AS total_out,
               SUM(qty_in) - SUM(qty_out) AS net_stock
        FROM combined
        GROUP BY project_code, item_code, batch_no, exp_date
        HAVING net_stock <> 0 OR SUM(qty_in) > 0
        ORDER BY project_code, item_code,
                 CASE WHEN exp_date='' OR exp_date IS NULL THEN 1 ELSE 0 END,
                 exp_date ASC
    '''
    return conn.execute(sql, params_r + params_b + params_i + params_o).fetchall()


@app.route('/api/reports/stock-summary', methods=['GET'])
@login_required
def rpt_stock_summary():
    conn = None
    try:
        conn    = _reports_db()
        project = request.args.get('project') or None
        item    = request.args.get('item') or None
        rows    = _stock_summary_rows(conn, project, item)
        return jsonify({'success': True, 'rows': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/reports/stock-summary/export', methods=['GET'])
@login_required
def rpt_stock_summary_export():
    conn = None
    try:
        conn    = _reports_db()
        project = request.args.get('project') or None
        item    = request.args.get('item') or None
        rows    = _stock_summary_rows(conn, project, item)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock Summary'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0

        fill   = PatternFill(start_color='1F3A8A', end_color='1F3A8A', fill_type='solid')
        hfont  = Font(bold=True, color='FFFFFF', size=10)
        headers = [
            ('Project', 14), ('Item Code', 14), ('Description', 34),
            ('Batch No', 14), ('Exp Date', 11),
            ('Total IN', 12), ('Total OUT', 12), ('Net Stock', 12),
        ]
        for c, (h, w) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hfont
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = w

        alt = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
        for i, r in enumerate(rows, 2):
            vals = [r['project_code'], r['item_code'], r['item_description'],
                    r['batch_no'], r['exp_date'],
                    r['total_in'], r['total_out'], r['net_stock']]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v)
                if i % 2 == 0:
                    cell.fill = alt

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"stock_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/reports/stock-card', methods=['GET'])
@login_required
def rpt_stock_card():
    """Full transaction history for a specific item + project (stock card)."""
    conn    = None
    item    = request.args.get('item', '').strip()
    project = request.args.get('project', '').strip()
    if not item:
        return jsonify({'success': False, 'message': 'item param required'}), 400
    try:
        conn = _reports_db()
        # Receptions
        rec_rows = conn.execute('''
            SELECT st.received_at AS txn_date, st.transaction_type AS doc_type,
                   st.reception_number AS document_number, st.project_code,
                   st.item_code, st.item_description, st.batch_no, st.exp_date,
                   st.qty_received AS qty_in, 0.0 AS qty_out,
                   'RECEPTION' AS source,
                   u.username AS user_name
            FROM stock_transactions st
            LEFT JOIN users u ON u.id = st.received_by
            WHERE st.item_code = ?
              AND (st.transaction_type = 'RECEPTION')
              AND (? = '' OR st.project_code = ?)
            ORDER BY st.received_at ASC
        ''', (item, project, project)).fetchall()

        # IN movements
        in_rows = conn.execute('''
            SELECT m.movement_date AS txn_date, m.doc_type,
                   m.document_number, m.dest_project AS project_code,
                   ml.item_code, ml.item_description, ml.batch_no, ml.exp_date,
                   ml.qty AS qty_in, 0.0 AS qty_out,
                   'IN' AS source,
                   u.username AS user_name
            FROM movement_lines ml
            JOIN movements m ON m.id = ml.movement_id
            LEFT JOIN users u ON u.id = m.created_by
            WHERE ml.item_code = ?
              AND m.movement_type = 'IN' AND m.status = 'Confirmed'
              AND (? = '' OR m.dest_project = ?)
            ORDER BY m.movement_date ASC
        ''', (item, project, project)).fetchall()

        # OUT movements
        out_rows = conn.execute('''
            SELECT m.movement_date AS txn_date, m.doc_type,
                   m.document_number, m.source_project AS project_code,
                   ml.item_code, ml.item_description, ml.batch_no, ml.exp_date,
                   0.0 AS qty_in, ml.qty AS qty_out,
                   'OUT' AS source,
                   u.username AS user_name
            FROM movement_lines ml
            JOIN movements m ON m.id = ml.movement_id
            LEFT JOIN users u ON u.id = m.created_by
            WHERE ml.item_code = ?
              AND m.movement_type = 'OUT' AND m.status = 'Confirmed'
              AND (? = '' OR m.source_project = ?)
            ORDER BY m.movement_date ASC
        ''', (item, project, project)).fetchall()

        all_txns = sorted(
            [dict(r) for r in (list(rec_rows) + list(in_rows) + list(out_rows))],
            key=lambda x: (x.get('txn_date') or '')
        )
        # Running balance
        balance = 0.0
        for t in all_txns:
            balance += (t['qty_in'] or 0) - (t['qty_out'] or 0)
            t['running_balance'] = round(balance, 4)

        return jsonify({'success': True, 'transactions': all_txns,
                        'item': item, 'project': project or 'ALL'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/reports/stock-card/export', methods=['GET'])
@login_required
def rpt_stock_card_export():
    item    = request.args.get('item', '').strip()
    project = request.args.get('project', '').strip()
    if not item:
        return jsonify({'success': False, 'message': 'item param required'}), 400
    conn = None
    try:
        conn     = _reports_db()
        resp     = rpt_stock_card()
        txns     = resp.get_json()['transactions']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock Card'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0

        ws['A1'] = f"Stock Card — {item}{' | Project: ' + project if project else ''}"
        ws['A1'].font = Font(bold=True, size=13, color='1F3A8A')
        ws.merge_cells('A1:L1')

        fill   = PatternFill(start_color='1F3A8A', end_color='1F3A8A', fill_type='solid')
        hfont  = Font(bold=True, color='FFFFFF', size=10)
        headers = [
            ('Date', 14), ('Type', 8), ('Document No', 18), ('Project', 12),
            ('Batch', 14), ('Exp Date', 11),
            ('IN Qty', 10), ('OUT Qty', 10), ('Balance', 10),
            ('User', 14), ('Source', 10),
        ]
        for c, (h, w) in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = hfont
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = w

        alt = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
        for i, t in enumerate(txns, 4):
            vals = [
                t.get('txn_date'), t.get('doc_type'), t.get('document_number'),
                t.get('project_code'), t.get('batch_no'), t.get('exp_date'),
                t.get('qty_in'), t.get('qty_out'), t.get('running_balance'),
                t.get('user_name'), t.get('source'),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v)
                if i % 2 == 0:
                    cell.fill = alt

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_item = _re.sub(r'[^\w]', '_', item)
        fname = f"stock_card_{safe_item}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/reports/transactions', methods=['GET'])
@login_required
def rpt_transactions():
    """All confirmed movements (IN + OUT) with filters."""
    conn    = None
    project = request.args.get('project') or None
    doc_type = request.args.get('doc_type') or None
    direction = request.args.get('direction') or None
    date_from = request.args.get('date_from') or None
    date_to   = request.args.get('date_to') or None
    page   = int(request.args.get('page', 1))
    limit  = int(request.args.get('limit', 100))
    offset = (page - 1) * limit
    try:
        conn = _reports_db()
        wheres = ["m.status = 'Confirmed'"]
        params = []
        if project:
            wheres.append("(m.source_project=? OR m.dest_project=?)")
            params += [project, project]
        if direction:
            wheres.append("m.movement_type=?")
            params.append(direction.upper())
        if doc_type:
            wheres.append("m.doc_type=?")
            params.append(doc_type)
        if date_from:
            wheres.append("m.movement_date >= ?")
            params.append(date_from)
        if date_to:
            wheres.append("m.movement_date <= ?")
            params.append(date_to)
        where_clause = "WHERE " + " AND ".join(wheres)

        rows = conn.execute(f'''
            SELECT m.id, m.document_number, m.movement_type, m.doc_type,
                   m.movement_date, m.source_project, m.dest_project,
                   m.total_weight_kg, m.total_volume_m3, m.notes, m.created_at,
                   u.username  AS created_by_name,
                   eu.name     AS end_user_name,
                   tp.name     AS third_party_name,
                   (SELECT COUNT(*) FROM movement_lines ml WHERE ml.movement_id=m.id) AS line_count
            FROM movements m
            LEFT JOIN users u      ON u.id = m.created_by
            LEFT JOIN end_users eu ON eu.end_user_id = m.end_user_id
            LEFT JOIN third_parties tp ON tp.third_party_id = m.third_party_id
            {where_clause}
            ORDER BY m.movement_date DESC, m.created_at DESC
            LIMIT ? OFFSET ?
        ''', params + [limit, offset]).fetchall()

        total = conn.execute(
            f"SELECT COUNT(*) FROM movements m {where_clause}", params
        ).fetchone()[0]

        return jsonify({'success': True, 'movements': [dict(r) for r in rows],
                        'total': total, 'page': page, 'limit': limit})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/reports/transactions/export', methods=['GET'])
@login_required
def rpt_transactions_export():
    conn = None
    try:
        conn      = _reports_db()
        project   = request.args.get('project') or None
        doc_type  = request.args.get('doc_type') or None
        direction = request.args.get('direction') or None
        date_from = request.args.get('date_from') or None
        date_to   = request.args.get('date_to') or None

        wheres = ["m.status = 'Confirmed'"]
        params = []
        if project:
            wheres.append("(m.source_project=? OR m.dest_project=?)")
            params += [project, project]
        if direction:
            wheres.append("m.movement_type=?"); params.append(direction.upper())
        if doc_type:
            wheres.append("m.doc_type=?");     params.append(doc_type)
        if date_from:
            wheres.append("m.movement_date>=?"); params.append(date_from)
        if date_to:
            wheres.append("m.movement_date<=?"); params.append(date_to)
        where_clause = "WHERE " + " AND ".join(wheres)

        rows = conn.execute(f'''
            SELECT m.document_number, m.movement_type, m.doc_type, m.movement_date,
                   m.source_project, m.dest_project,
                   m.total_weight_kg, m.total_volume_m3, m.notes,
                   u.username AS created_by_name,
                   eu.name AS end_user_name, tp.name AS third_party_name,
                   ml.line_no, ml.item_code, ml.item_description,
                   ml.batch_no, ml.exp_date, ml.qty, ml.unit,
                   ml.unit_price, ml.currency, ml.total_value,
                   ml.weight_kg, ml.volume_m3
            FROM movements m
            LEFT JOIN users u      ON u.id = m.created_by
            LEFT JOIN end_users eu ON eu.end_user_id = m.end_user_id
            LEFT JOIN third_parties tp ON tp.third_party_id = m.third_party_id
            LEFT JOIN movement_lines ml ON ml.movement_id = m.id
            {where_clause}
            ORDER BY m.movement_date DESC, m.id, ml.line_no
        ''', params).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transactions'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0

        fill  = PatternFill(start_color='1F3A8A', end_color='1F3A8A', fill_type='solid')
        hfnt  = Font(bold=True, color='FFFFFF', size=9)
        headers = [
            'Document No', 'Direction', 'Type', 'Date', 'From Project', 'To Project',
            'End User', 'Third Party', 'Line', 'Item Code', 'Description',
            'Batch', 'Exp Date', 'Qty', 'Unit',
            'Unit Price', 'Currency', 'Total Value',
            'Weight kg', 'Volume m3', 'Notes',
        ]
        widths = [18, 8, 8, 12, 14, 14, 20, 20, 5, 14, 32,
                  14, 11, 8, 8, 11, 10, 13, 11, 11, 24]
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hfnt
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = w

        alt = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
        for i, r in enumerate(rows, 2):
            vals = [
                r['document_number'], r['movement_type'], r['doc_type'], r['movement_date'],
                r['source_project'], r['dest_project'],
                r['end_user_name'], r['third_party_name'],
                r['line_no'], r['item_code'], r['item_description'],
                r['batch_no'], r['exp_date'], r['qty'], r['unit'],
                r['unit_price'], r['currency'], r['total_value'],
                r['weight_kg'], r['volume_m3'], r['notes'],
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v)
                if i % 2 == 0:
                    cell.fill = alt

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Expiry Report ─────────────────────────────────────────────────────────

@app.route('/api/reports/expiry', methods=['GET'])
@login_required
def rpt_expiry():
    """Items with an expiry date within the next N days (or already expired)."""
    conn = None
    try:
        conn        = _reports_db()
        project     = request.args.get('project') or None
        within_days = int(request.args.get('within_days') or 90)
        rows        = _stock_summary_rows(conn, project, None)
        today       = datetime.now().date()
        result      = []
        for r in rows:
            if not r['exp_date']:
                continue
            try:
                exp_date = datetime.strptime(r['exp_date'], '%Y-%m-%d').date()
            except Exception:
                continue
            days_left = (exp_date - today).days
            if days_left > within_days:
                continue
            if (r['net_stock'] or 0) <= 0:
                continue
            if days_left < 0:
                status = 'Expired'
            elif days_left < 30:
                status = 'Critical'
            elif days_left < 90:
                status = 'Warning'
            else:
                status = 'OK'
            result.append({
                'project_code':    r['project_code'],
                'item_code':       r['item_code'],
                'item_description':r['item_description'],
                'batch_no':        r['batch_no'],
                'exp_date':        r['exp_date'],
                'days_left':       days_left,
                'net_stock':       r['net_stock'],
                'status':          status,
            })
        result.sort(key=lambda x: x['days_left'])
        return jsonify({'success': True, 'rows': result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/reports/expiry/export', methods=['GET'])
@login_required
def rpt_expiry_export():
    """Export expiry report to Excel."""
    conn = None
    try:
        conn        = _reports_db()
        project     = request.args.get('project') or None
        within_days = int(request.args.get('within_days') or 90)
        rows        = _stock_summary_rows(conn, project, None)
        today       = datetime.now().date()
        result      = []
        for r in rows:
            if not r['exp_date']:
                continue
            try:
                exp_date = datetime.strptime(r['exp_date'], '%Y-%m-%d').date()
            except Exception:
                continue
            days_left = (exp_date - today).days
            if days_left > within_days:
                continue
            if (r['net_stock'] or 0) <= 0:
                continue
            status = 'Expired' if days_left < 0 else ('Critical' if days_left < 30 else ('Warning' if days_left < 90 else 'OK'))
            result.append({
                'project_code':    r['project_code'],
                'item_code':       r['item_code'],
                'item_description':r['item_description'],
                'batch_no':        r['batch_no'],
                'exp_date':        r['exp_date'],
                'days_left':       days_left,
                'net_stock':       r['net_stock'],
                'status':          status,
            })
        result.sort(key=lambda x: x['days_left'])

        # Get mission name for heading
        mission_row = conn.execute(
            "SELECT mission_name, mission_abbreviation FROM mission_details WHERE is_active=1 LIMIT 1"
        ).fetchone()
        mission_name  = mission_row['mission_name']        if mission_row else ''
        mission_abbr  = mission_row['mission_abbreviation'] if mission_row else ''

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Expiry Report'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage  = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        NUM_COLS = 8

        # ── Title block (rows 1-4) ─────────────────────────────────────────
        title_fill = PatternFill(start_color='1F3A8A', end_color='1F3A8A', fill_type='solid')
        title_font = Font(bold=True, color='FFFFFF', size=13)
        sub_font   = Font(bold=False, color='FFFFFF', size=10)
        white_bold = Font(bold=True, color='FFFFFF', size=10)

        # Row 1: Mission name (left) + report title (right)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        c1 = ws.cell(row=1, column=1, value=mission_name or mission_abbr or 'EXPIRY REPORT')
        c1.font = title_font
        c1.fill = title_fill
        c1.alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=NUM_COLS)
        c2 = ws.cell(row=1, column=5, value='EXPIRY REPORT')
        c2.font = title_font
        c2.fill = title_fill
        c2.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[1].height = 22

        # Row 2: Project filter + within days
        proj_label = f"Project: {project}" if project else 'Project: All'
        days_label = f"Expiring within: {within_days} days" if within_days < 9999 else 'All expired + expiring'
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        c3 = ws.cell(row=2, column=1, value=proj_label)
        c3.font = sub_font; c3.fill = title_fill
        c3.alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=NUM_COLS)
        c4 = ws.cell(row=2, column=5, value=days_label)
        c4.font = sub_font; c4.fill = title_fill
        c4.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[2].height = 16

        # Row 3: Generated date + summary counts
        expired_cnt  = sum(1 for r in result if r['status']=='Expired')
        critical_cnt = sum(1 for r in result if r['status']=='Critical')
        warning_cnt  = sum(1 for r in result if r['status']=='Warning')
        gen_label = f"Generated: {today.strftime('%Y-%m-%d')}    |    Expired: {expired_cnt}  Critical: {critical_cnt}  Warning: {warning_cnt}  Total: {len(result)}"
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NUM_COLS)
        c5 = ws.cell(row=3, column=1, value=gen_label)
        c5.font = white_bold; c5.fill = title_fill
        c5.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[3].height = 16

        # Row 4: blank spacer with title fill
        for col in range(1, NUM_COLS + 1):
            ws.cell(row=4, column=col).fill = title_fill
        ws.row_dimensions[4].height = 6

        # ── Column headers (row 5) ────────────────────────────────────────
        headers = [('Project', 12), ('Item Code', 16), ('Description', 34),
                   ('Batch', 14), ('Exp Date', 11), ('Days Left', 10),
                   ('Net Stock', 11), ('Status', 10)]
        hfnt = Font(bold=True, color='FFFFFF')
        hfill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
        thin  = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c, (h, w) in enumerate(headers, 1):
            cell = ws.cell(row=5, column=c, value=h)
            cell.font   = hfnt
            cell.fill   = hfill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[5].height = 16

        # ── Data rows (starting row 6) ────────────────────────────────────
        red_fill    = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        orange_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        for i, r in enumerate(result, 6):
            vals = [r['project_code'], r['item_code'], r['item_description'],
                    r['batch_no'], r['exp_date'], r['days_left'],
                    round(r['net_stock'], 3), r['status']]
            row_fill = red_fill if r['days_left'] < 30 else (orange_fill if r['days_left'] < 90 else None)
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v)
                cell.border = border
                if row_fill:
                    cell.fill = row_fill

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"expiry_report_{today.strftime('%Y%m%d')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════
#  INVENTORY  (Phase 5)
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/api/inventory/parcels', methods=['GET'])
@login_required
def inv_parcels():
    """Received parcels grouped by parcel_number for physical inventory."""
    project = request.args.get('project') or None
    conn    = None
    try:
        conn = _mov_db()
        where = "WHERE reception_number IS NOT NULL"
        params = []
        if project:
            where += " AND project_code = ?"
            params.append(project)
        rows = conn.execute(f'''
            SELECT parcel_number, packing_ref, project_code, pallet_number,
                   MAX(order_type) AS order_type,
                   SUM(weight_kg)  AS total_weight,
                   SUM(volume_m3)  AS total_volume,
                   COUNT(*)        AS item_count,
                   MAX(received_at) AS received_at
            FROM basic_data
            {where}
            GROUP BY parcel_number
            ORDER BY project_code, parcel_number
        ''', params).fetchall()
        return jsonify({'success': True, 'parcels': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/inventory/items', methods=['GET'])
@login_required
def inv_items():
    """System stock by project for item-level physical count."""
    project = request.args.get('project') or None
    conn    = None
    try:
        conn  = _mov_db()
        rows  = _stock_summary_rows(conn, project, None)
        items = []
        for r in rows:
            if (r['net_stock'] or 0) <= 0:
                continue
            d = dict(r)
            # Look up parcel_number and pallet_number from basic_data
            bd = conn.execute(
                '''SELECT parcel_number, pallet_number FROM basic_data
                   WHERE item_code=? AND (project_code=? OR (project_code IS NULL AND ? IS NULL))
                     AND (batch_no_received=? OR batch_no=?)
                   LIMIT 1''',
                (d.get('item_code'), d.get('project_code'), d.get('project_code'),
                 d.get('batch_no') or '', d.get('batch_no') or '')
            ).fetchone()
            d['parcel_number'] = bd['parcel_number'] if bd else None
            d['pallet_number'] = bd['pallet_number'] if bd else None
            items.append(d)
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/inventory/count', methods=['POST'])
@login_required
def inv_save_count():
    conn = None
    try:
        data       = request.get_json()
        count_date = data.get('count_date') or datetime.now().strftime('%Y-%m-%d')
        project    = data.get('project_code')
        count_type = data.get('count_type', 'item')
        notes      = data.get('notes')
        lines      = data.get('lines', [])
        conn       = _mov_db()

        cur = conn.execute('''
            INSERT INTO inventory_counts (count_date, project_code, count_type, notes, created_by)
            VALUES (?, ?, ?, ?, ?)
        ''', (count_date, project, count_type, notes, current_user.id))
        count_id = cur.lastrowid

        for ln in lines:
            system_qty   = float(ln.get('system_qty')   or 0)
            physical_qty = float(ln.get('physical_qty') or 0)
            variance     = physical_qty - system_qty
            conn.execute('''
                INSERT INTO inventory_count_lines
                    (count_id, parcel_number, item_code, item_description,
                     batch_no, exp_date, system_qty, physical_qty, variance, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (count_id, ln.get('parcel_number'), ln.get('item_code'),
                  ln.get('item_description'), ln.get('batch_no'), ln.get('exp_date'),
                  system_qty, physical_qty, variance, ln.get('notes')))

        conn.commit()
        return jsonify({'success': True, 'count_id': count_id})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/inventory/counts', methods=['GET'])
@login_required
def inv_counts_list():
    conn = None
    try:
        conn = _mov_db()
        rows = conn.execute('''
            SELECT ic.*, u.username AS created_by_name,
                   (SELECT COUNT(*) FROM inventory_count_lines l WHERE l.count_id=ic.id) AS line_count,
                   (SELECT SUM(ABS(l.variance)) FROM inventory_count_lines l WHERE l.count_id=ic.id) AS total_variance
            FROM inventory_counts ic
            LEFT JOIN users u ON u.id = ic.created_by
            ORDER BY ic.created_at DESC
        ''').fetchall()
        return jsonify({'success': True, 'counts': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/inventory/counts/<int:count_id>', methods=['GET'])
@login_required
def inv_count_detail(count_id):
    conn = None
    try:
        conn = _mov_db()
        hdr  = conn.execute(
            "SELECT ic.*, u.username AS created_by_name "
            "FROM inventory_counts ic LEFT JOIN users u ON u.id=ic.created_by "
            "WHERE ic.id=?", (count_id,)
        ).fetchone()
        if not hdr:
            return jsonify({'success': False, 'message': 'Not found'}), 404
        lines = conn.execute(
            "SELECT * FROM inventory_count_lines WHERE count_id=? ORDER BY id",
            (count_id,)
        ).fetchall()
        return jsonify({'success': True, 'count': dict(hdr),
                        'lines': [dict(l) for l in lines]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


@app.route('/api/inventory/counts/<int:count_id>/export', methods=['GET'])
@login_required
def inv_count_export(count_id):
    conn = None
    try:
        conn  = _mov_db()
        hdr   = conn.execute(
            "SELECT ic.*, u.username AS created_by_name "
            "FROM inventory_counts ic LEFT JOIN users u ON u.id=ic.created_by "
            "WHERE ic.id=?", (count_id,)
        ).fetchone()
        if not hdr:
            return jsonify({'success': False, 'message': 'Not found'}), 404
        lines = conn.execute(
            "SELECT * FROM inventory_count_lines WHERE count_id=? ORDER BY id",
            (count_id,)
        ).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventory Count'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0

        ws['A1'] = f"Inventory Count — {hdr['count_type'].upper()}"
        ws['A1'].font = Font(bold=True, size=13, color='1F3A8A')
        ws.merge_cells('A1:K1')
        ws['A2'] = f"Date: {hdr['count_date']}  |  Project: {hdr['project_code'] or 'ALL'}  |  By: {hdr['created_by_name']}"
        ws.merge_cells('A2:K2')

        fill   = PatternFill(start_color='1F3A8A', end_color='1F3A8A', fill_type='solid')
        hfont  = Font(bold=True, color='FFFFFF', size=10)
        is_parcel = (hdr['count_type'] == 'parcel')
        if is_parcel:
            headers = [('Parcel No', 16), ('Item Code', 14), ('Description', 34),
                       ('System Qty', 12), ('Physical Qty', 14), ('Variance', 12), ('Notes', 24)]
        else:
            headers = [('Item Code', 14), ('Description', 34), ('Batch No', 14), ('Exp Date', 11),
                       ('System Qty', 12), ('Physical Qty', 14), ('Variance', 12), ('Notes', 24)]
        for c, (h, w) in enumerate(headers, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = hfont
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = w

        RED   = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        GREEN = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        ALT   = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')

        for i, ln in enumerate(lines, 5):
            if is_parcel:
                vals = [ln['parcel_number'], ln['item_code'], ln['item_description'],
                        ln['system_qty'], ln['physical_qty'], ln['variance'], ln['notes']]
            else:
                vals = [ln['item_code'], ln['item_description'],
                        ln['batch_no'], ln['exp_date'],
                        ln['system_qty'], ln['physical_qty'], ln['variance'], ln['notes']]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v)
                var = ln['variance'] or 0
                if var < 0:
                    cell.fill = RED
                elif var > 0:
                    cell.fill = GREEN
                elif i % 2 == 0:
                    cell.fill = ALT

        ws.freeze_panes = 'A5'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"inventory_count_{count_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ── Blank count sheet export (for printing / field use) ──────────────────
@app.route('/api/inventory/blank-sheet', methods=['GET'])
@login_required
def inv_blank_sheet():
    """Export blank count sheet pre-filled with system quantities for a project."""
    project    = request.args.get('project') or None
    count_type = request.args.get('type', 'item')
    conn = None
    try:
        conn = _mov_db()
        wb   = openpyxl.Workbook()
        ws   = wb.active
        ws.title = 'Count Sheet'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0

        ws['A1'] = f"Physical Inventory Count Sheet — {count_type.upper()}"
        ws['A1'].font = Font(bold=True, size=13)
        ws.merge_cells('A1:I1')
        ws['A2'] = f"Project: {project or 'ALL'}  |  Date: _______________  |  Counted by: _______________"
        ws.merge_cells('A2:I2')

        fill  = PatternFill(start_color='1F3A8A', end_color='1F3A8A', fill_type='solid')
        hfont = Font(bold=True, color='FFFFFF', size=10)

        if count_type == 'parcel':
            rows = conn.execute('''
                SELECT parcel_number, packing_ref, project_code, pallet_number,
                       COUNT(*) AS item_count, SUM(weight_kg) AS total_weight
                FROM basic_data
                WHERE reception_number IS NOT NULL
                  AND (? IS NULL OR project_code=?)
                GROUP BY parcel_number
                ORDER BY project_code, parcel_number
            ''', (project, project)).fetchall()

            headers = [('Parcel No', 16), ('Packing Ref', 14), ('Project', 12),
                       ('Pallet', 12), ('Items', 8), ('Weight', 10),
                       ('Physical Count', 16), ('OK?', 6), ('Notes', 24)]
            for c, (h, w) in enumerate(headers, 1):
                cell = ws.cell(row=4, column=c, value=h)
                cell.font = hfont; cell.fill = fill
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[cell.column_letter].width = w
            for i, r in enumerate(rows, 5):
                ws.cell(row=i, column=1, value=r['parcel_number'])
                ws.cell(row=i, column=2, value=r['packing_ref'])
                ws.cell(row=i, column=3, value=r['project_code'])
                ws.cell(row=i, column=4, value=r['pallet_number'])
                ws.cell(row=i, column=5, value=r['item_count'])
                ws.cell(row=i, column=6, value=r['total_weight'])
        else:
            stock_rows = _stock_summary_rows(conn, project, None)
            stock_rows = [r for r in stock_rows if (r['net_stock'] or 0) > 0]

            headers = [('Item Code', 14), ('Description', 34), ('Batch', 14), ('Exp Date', 11),
                       ('Project', 12), ('System Qty', 12), ('Physical Qty', 14), ('Variance', 12), ('Notes', 24)]
            for c, (h, w) in enumerate(headers, 1):
                cell = ws.cell(row=4, column=c, value=h)
                cell.font = hfont; cell.fill = fill
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[cell.column_letter].width = w
            for i, r in enumerate(stock_rows, 5):
                ws.cell(row=i, column=1, value=r['item_code'])
                ws.cell(row=i, column=2, value=r['item_description'])
                ws.cell(row=i, column=3, value=r['batch_no'])
                ws.cell(row=i, column=4, value=r['exp_date'])
                ws.cell(row=i, column=5, value=r['project_code'])
                ws.cell(row=i, column=6, value=r['net_stock'])

        ws.freeze_panes = 'A5'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"count_sheet_{count_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════
#  PALLET CHANGE  (Phase 6)
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/api/cargo/change-pallet', methods=['PATCH'])
@login_required
def cr_change_pallet():
    conn = None
    try:
        data           = request.get_json()
        parcel_number  = (data.get('parcel_number') or '').strip()
        new_pallet     = (data.get('new_pallet')    or '').strip()
        if not parcel_number or not new_pallet:
            return jsonify({'success': False,
                            'message': 'parcel_number and new_pallet are required'}), 400
        conn = _mov_db()
        conn.execute(
            "UPDATE basic_data SET pallet_number=? WHERE parcel_number=?",
            (new_pallet, parcel_number))
        conn.execute(
            "UPDATE stock_transactions SET pallet_number=? WHERE parcel_number=?",
            (new_pallet, parcel_number))
        conn.commit()
        return jsonify({'success': True, 'new_pallet': new_pallet})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn: conn.close()


if __name__ == '__main__':
    app.run(debug=True, port=5000)