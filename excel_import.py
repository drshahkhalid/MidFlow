import openpyxl
from datetime import datetime
import uuid
from database import get_db_connection
from utils import normalize_number, normalize_date

class ExcelImporter:
    """
    Flexible Excel importer that handles column mapping
    """
    
    # Default column mappings - can be customized
    DEFAULT_MAPPINGS = {
        'file1': {  # Packing list file
            'Packing ref': 'packing_ref',
            'Line no': 'line_no',
            'Item code': 'item_code',
            'Item description': 'item_description',
            'Qty unit. tot.': 'qty_unit_tot',
            'Packaging': 'packaging',
            'Parcel n°': 'parcel_no',
            'Nb parcels': 'nb_parcels',
            'Batch no': 'batch_no',
            'Exp. date': 'exp_date',
            'Kg (total)': 'kg_total',
            'dm3 (total)': 'dm3_total',
        },
        'file2': {  # Reception file
            'Goods reception': 'packing_ref',  # This matches with Packing ref
            'Transport reception': 'transport_reception',
            'Sub folder': 'sub_folder',
            'Field ref.': 'field_ref',
            'Ref op MSFL': 'ref_op_msfl',
            'Parcel nb': 'parcel_nb',
            'Weight (kg)': 'weight_kg',
            'Volume (m3)': 'volume_m3',
            'Invoice/credit note ref': 'invoice_credit_note_ref',
            'Estim. value (for items) (eu)': 'estim_value_eu',
        }
    }
    
    @staticmethod
    def generate_unique_id():
        """Generate a unique ID for each record"""
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        short_uuid = str(uuid.uuid4())[:8]
        return f"BD-{timestamp}-{short_uuid}"
    
    @staticmethod
    def detect_header_row(worksheet, max_rows=10):
        """
        Detect which row contains headers by looking for text-heavy rows
        """
        for row_idx in range(1, max_rows + 1):
            row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            # Count non-empty cells
            non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
            if non_empty >= 3:  # At least 3 column headers
                return row_idx, row
        return 1, list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    @staticmethod
    def normalize_header(header):
        """Normalize header names for flexible matching"""
        if header is None:
            return ""
        return str(header).strip().lower().replace('  ', ' ')
    
    @staticmethod
    def find_column_index(headers, possible_names):
        """
        Find column index by trying multiple possible names
        """
        normalized_headers = [ExcelImporter.normalize_header(h) for h in headers]
        
        for name in possible_names:
            normalized_name = ExcelImporter.normalize_header(name)
            if normalized_name in normalized_headers:
                return normalized_headers.index(normalized_name)
        return None
    
    @staticmethod
    def read_excel_file(file_path, file_type='file1', custom_mapping=None):
        """
        Read Excel file and return data with flexible column mapping
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Detect header row
        header_row_idx, headers = ExcelImporter.detect_header_row(ws)
        
        # Use custom mapping or default
        mapping = custom_mapping or ExcelImporter.DEFAULT_MAPPINGS.get(file_type, {})
        
        # Build column index map
        column_indices = {}
        for source_col, target_col in mapping.items():
            idx = ExcelImporter.find_column_index(headers, [source_col])
            if idx is not None:
                column_indices[target_col] = idx
        
        # Read data rows
        data_rows = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            
            row_data = {}
            for target_col, col_idx in column_indices.items():
                value = row[col_idx] if col_idx < len(row) else None
                
                # Normalize data types
                if value is not None:
                    if target_col in ['qty_unit_tot', 'kg_total', 'dm3_total', 'weight_kg', 'volume_m3', 'estim_value_eu']:
                        value = normalize_number(value)
                    elif target_col in ['exp_date']:
                        value = normalize_date(value)
                    elif target_col in ['nb_parcels']:
                        try:
                            value = int(normalize_number(value) or 0)
                        except:
                            value = 0
                    else:
                        value = str(value).strip() if value else None
                
                row_data[target_col] = value
            
            if row_data:  # Only add if we got some data
                data_rows.append(row_data)
        
        wb.close()
        return data_rows, list(column_indices.keys())
    
    @staticmethod
    def merge_data(file1_data, file2_data, match_column='packing_ref'):
        """
        Merge data from two files based on matching column
        """
        # Create lookup dictionary for file2 data
        file2_lookup = {}
        for row in file2_data:
            key = row.get(match_column)
            if key:
                if key not in file2_lookup:
                    file2_lookup[key] = []
                file2_lookup[key].append(row)
        
        # Merge data
        merged_data = []
        for row1 in file1_data:
            match_key = row1.get(match_column)
            
            if match_key and match_key in file2_lookup:
                # Merge with all matching rows from file2
                for row2 in file2_lookup[match_key]:
                    merged_row = {**row1, **row2}
                    merged_data.append(merged_row)
            else:
                # No match, add file1 data only
                merged_data.append(row1)
        
        return merged_data
    
    @staticmethod
    def import_to_database(data_rows, source_file, user_id):
        """
        Import merged data into database
        """
        conn = get_db_connection()
        cursor = conn.cursor()
        
        imported_count = 0
        errors = []
        
        for row in data_rows:
            try:
                unique_id = ExcelImporter.generate_unique_id()
                
                cursor.execute('''
                    INSERT INTO basic_data (
                        unique_id, packing_ref, line_no, item_code, item_description,
                        qty_unit_tot, packaging, parcel_no, nb_parcels, batch_no,
                        exp_date, kg_total, dm3_total, transport_reception, sub_folder,
                        field_ref, ref_op_msfl, parcel_nb, weight_kg, volume_m3,
                        invoice_credit_note_ref, estim_value_eu, source_file, imported_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    unique_id,
                    row.get('packing_ref'),
                    row.get('line_no'),
                    row.get('item_code'),
                    row.get('item_description'),
                    row.get('qty_unit_tot'),
                    row.get('packaging'),
                    row.get('parcel_no'),
                    row.get('nb_parcels'),
                    row.get('batch_no'),
                    row.get('exp_date'),
                    row.get('kg_total'),
                    row.get('dm3_total'),
                    row.get('transport_reception'),
                    row.get('sub_folder'),
                    row.get('field_ref'),
                    row.get('ref_op_msfl'),
                    row.get('parcel_nb'),
                    row.get('weight_kg'),
                    row.get('volume_m3'),
                    row.get('invoice_credit_note_ref'),
                    row.get('estim_value_eu'),
                    source_file,
                    user_id
                ))
                imported_count += 1
            except Exception as e:
                errors.append(f"Row error: {str(e)}")
        
        conn.commit()
        conn.close()
        
        return imported_count, errors

# Example usage
if __name__ == '__main__':
    # Test import
    importer = ExcelImporter()
    print("Excel Importer ready!")